"""AITable API 调试工具 - 浏览器版

零依赖，纯 Python 标准库（http.server + 内嵌 HTML/JS）。
运行：python3 aitable_web_tool.py
"""

import csv
import io
import json
import os
import threading
import time
import uuid
import webbrowser
import urllib.request
import urllib.error
from datetime import datetime
from http.server import BaseHTTPRequestHandler, HTTPServer

PORT = 8765
HERE = os.path.dirname(os.path.abspath(__file__))
SCHEDULES_FILE = os.path.join(HERE, "aitable_schedules.json")
LOGS_FILE = os.path.join(HERE, "aitable_schedule_logs.json")
LOG_LIMIT = 200

INDEX_HTML = r"""<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="UTF-8">
<title>AITable API 调试工具</title>
<!-- SheetJS：解析 Excel/CSV。失败也不影响其他功能 -->
<script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js"
        onerror="window._sheetjsLoadFailed=true;"></script>
<style>
  /* ---------- 设计系统 ---------- */
  :root {
    --bg: #f8fafc;
    --card: #ffffff;
    --border: #e5e7eb;
    --border-strong: #d1d5db;
    --text: #111827;
    --text-2: #4b5563;
    --text-3: #9ca3af;
    --primary: #6366f1;
    --primary-hover: #4f46e5;
    --primary-soft: #eef2ff;
    --primary-border: #c7d2fe;
    --danger: #dc2626;
    --danger-soft: #fee2e2;
    --success: #059669;
    --warn-bg: #fffbeb;
    --warn-border: #fde68a;
    --warn-text: #78350f;
    --shadow: 0 1px 2px rgba(15,23,42,0.04), 0 1px 3px rgba(15,23,42,0.05);
    --shadow-lg: 0 4px 12px rgba(15,23,42,0.08);
    --radius: 10px;
    --radius-sm: 6px;
  }

  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, "Inter", "PingFang SC",
                 "Microsoft YaHei", system-ui, sans-serif;
    background: var(--bg);
    color: var(--text);
    font-size: 14px;
    line-height: 1.5;
    padding: 24px 16px 60px;
  }
  .container { max-width: 960px; margin: 0 auto; }

  /* ---------- 顶部 ---------- */
  .top-bar {
    display: flex; justify-content: space-between; align-items: center;
    margin-bottom: 20px; gap: 12px; flex-wrap: wrap;
  }
  .top-bar h1 {
    font-size: 20px; font-weight: 600; margin: 0;
    display: flex; align-items: center; gap: 10px;
  }
  .top-bar h1 .icon {
    width: 32px; height: 32px; border-radius: 8px;
    background: linear-gradient(135deg, #6366f1, #8b5cf6);
    display: inline-flex; align-items: center; justify-content: center;
    color: white; font-weight: bold; font-size: 16px;
  }
  .doc-link {
    color: var(--primary); text-decoration: none; font-size: 13px;
    padding: 7px 14px; border: 1px solid var(--primary-border);
    border-radius: var(--radius-sm); background: var(--primary-soft);
    display: inline-flex; align-items: center; gap: 6px;
    transition: all 0.15s;
  }
  .doc-link:hover { background: var(--primary); color: white; border-color: var(--primary); }

  /* ---------- 卡片 ---------- */
  .card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 18px 22px;
    margin-bottom: 14px;
    box-shadow: var(--shadow);
  }
  .card-head {
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 14px; gap: 10px;
  }
  .card-head h2 {
    font-size: 13px; font-weight: 600; color: var(--text-2);
    margin: 0; text-transform: uppercase; letter-spacing: 0.5px;
    display: flex; align-items: center; gap: 6px;
  }
  .card-head h2::before {
    content: ""; width: 3px; height: 14px; background: var(--primary);
    border-radius: 2px;
  }
  .card-head .meta {
    font-size: 12px; color: var(--text-3); font-weight: normal;
  }

  /* ---------- 表单字段（纵向） ---------- */
  .field {
    display: flex; flex-direction: column; gap: 6px;
  }
  .field-label {
    display: flex; align-items: center; gap: 6px;
    font-size: 13px; font-weight: 500; color: var(--text);
  }
  .field-label .api-name {
    color: var(--text-3); font-weight: 400; font-size: 11px;
    font-family: ui-monospace, "SF Mono", Menlo, monospace;
    background: #f3f4f6; padding: 1px 6px; border-radius: 3px;
  }
  .field-label .req {
    color: var(--danger); font-weight: bold;
  }
  .help-icon {
    display: inline-flex; align-items: center; justify-content: center;
    width: 16px; height: 16px; border-radius: 50%;
    background: #e5e7eb; color: var(--text-2); font-size: 11px;
    cursor: pointer; user-select: none; font-weight: 600;
    transition: all 0.15s;
  }
  .help-icon:hover { background: var(--primary); color: white; }
  .help-panel {
    padding: 10px 14px; background: var(--warn-bg);
    border: 1px solid var(--warn-border); border-radius: var(--radius-sm);
    font-size: 12px; color: var(--warn-text); line-height: 1.6;
    display: none;
  }
  .help-panel.open { display: block; }
  .help-panel::before {
    content: "📘 "; opacity: 0.7;
  }
  .field-hint {
    font-size: 12px; color: var(--text-3);
  }

  input[type="text"], input[type="password"], select, textarea {
    width: 100%;
    padding: 8px 12px; border: 1px solid var(--border-strong);
    border-radius: var(--radius-sm); font-size: 13px;
    font-family: inherit; background: white; color: var(--text);
    transition: border-color 0.15s, box-shadow 0.15s;
  }
  input:focus, select:focus, textarea:focus {
    outline: none; border-color: var(--primary);
    box-shadow: 0 0 0 3px rgba(99,102,241,0.15);
  }
  textarea {
    font-family: ui-monospace, "SF Mono", Menlo, monospace;
    resize: vertical; line-height: 1.5;
  }

  /* ---------- 双栏栅格 ---------- */
  .grid-2 {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
    gap: 14px;
  }
  .grid-2 .field.full { grid-column: 1 / -1; }

  /* ---------- 按钮 ---------- */
  button {
    background: var(--primary); color: white; border: none;
    padding: 8px 16px; border-radius: var(--radius-sm);
    cursor: pointer; font-size: 13px; font-weight: 500;
    transition: all 0.15s; font-family: inherit;
  }
  button:hover { background: var(--primary-hover); }
  button.secondary {
    background: white; color: var(--text);
    border: 1px solid var(--border-strong);
  }
  button.secondary:hover { background: #f3f4f6; }
  button.danger {
    background: var(--danger-soft); color: var(--danger); border: none;
  }
  button.danger:hover { background: #fecaca; }
  button.ghost {
    background: transparent; color: var(--text-2);
    padding: 6px 10px; font-size: 12px;
  }
  button.ghost:hover { background: #f3f4f6; color: var(--text); }
  button.small { padding: 5px 12px; font-size: 12px; }
  button.send {
    padding: 10px 24px; font-size: 14px;
    box-shadow: 0 2px 6px rgba(99,102,241,0.3);
  }
  .actions { display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }

  /* ---------- 模板区 ---------- */
  .template-bar {
    display: flex; gap: 8px; align-items: center; flex-wrap: wrap;
  }
  .template-bar select {
    flex: 1; min-width: 220px;
  }

  /* ---------- 记录编辑器 ---------- */
  .records-toolbar {
    display: flex; gap: 4px; margin-bottom: 14px;
    background: #f3f4f6; padding: 3px; border-radius: var(--radius-sm);
    width: fit-content;
  }
  .records-toolbar button {
    background: transparent; color: var(--text-2);
    padding: 5px 14px; border-radius: 4px; font-size: 12px;
  }
  .records-toolbar button.active {
    background: white; color: var(--primary);
    box-shadow: 0 1px 2px rgba(0,0,0,0.06);
  }
  .record-card {
    border: 1px solid var(--border); border-radius: var(--radius-sm);
    padding: 14px; margin-bottom: 12px; background: #fafafa;
  }
  .record-card-header {
    display: flex; justify-content: space-between; align-items: center;
    margin-bottom: 12px; padding-bottom: 10px;
    border-bottom: 1px dashed var(--border-strong);
  }
  .record-title {
    font-weight: 600; color: var(--primary); font-size: 13px;
    display: inline-flex; align-items: center; gap: 6px;
  }
  .record-title::before {
    content: "📝"; opacity: 0.6;
  }
  .recordid-row {
    display: flex; gap: 8px; align-items: center; margin-bottom: 10px;
  }
  .recordid-row label {
    font-size: 12px; color: var(--text-2); min-width: 60px;
  }
  .field-row {
    display: grid; grid-template-columns: 1fr 1.5fr 90px 32px;
    gap: 8px; margin-bottom: 8px; align-items: center;
  }
  .field-row input, .field-row select {
    padding: 6px 10px; font-size: 13px;
  }
  .field-row .del-btn {
    background: transparent; color: var(--text-3);
    border: none; cursor: pointer; padding: 4px;
    border-radius: 4px; font-size: 14px;
  }
  .field-row .del-btn:hover { background: var(--danger-soft); color: var(--danger); }
  .field-row.duplicate input {
    border-color: var(--danger) !important;
    background: #fee2e2 !important;
  }
  .duplicate-warning {
    background: var(--warn-bg); color: var(--warn-text);
    border: 1px solid var(--warn-border);
    padding: 10px 12px; border-radius: 4px;
    font-size: 12px; margin-bottom: 10px; line-height: 1.6;
  }
  .duplicate-warning b { color: var(--danger); }
  .field-row-head {
    display: grid; grid-template-columns: 1fr 1.5fr 90px 32px;
    gap: 8px; padding: 0 4px; margin-bottom: 6px;
    color: var(--text-3); font-size: 11px; font-weight: 500;
    text-transform: uppercase; letter-spacing: 0.3px;
  }

  /* ---------- 结果区 ---------- */
  .result-box {
    background: #0f172a; color: #e2e8f0; padding: 14px 16px;
    border-radius: var(--radius-sm);
    font-family: ui-monospace, "SF Mono", Menlo, monospace;
    font-size: 12px; white-space: pre-wrap; word-break: break-all;
    max-height: 500px; overflow-y: auto; min-height: 60px;
    line-height: 1.6;
  }
  .result-box.error { background: #450a0a; color: #fecaca; }
  .result-box:empty::before {
    content: "等待执行…"; color: #475569; font-style: italic;
  }

  /* ---------- 响应工具栏 + 表格视图 ---------- */
  .response-toolbar {
    display: flex; justify-content: space-between; align-items: center;
    margin-bottom: 10px; gap: 12px; flex-wrap: wrap;
  }
  .view-toggle {
    display: flex; gap: 3px; background: #f3f4f6;
    padding: 3px; border-radius: var(--radius-sm); width: fit-content;
  }
  .view-toggle button {
    background: transparent; color: var(--text-2);
    padding: 5px 14px; border-radius: 4px; font-size: 12px;
  }
  .view-toggle button.active {
    background: white; color: var(--primary);
    box-shadow: 0 1px 2px rgba(0,0,0,0.06);
  }
  .table-view {
    border: 1px solid var(--border); border-radius: var(--radius-sm);
    overflow-x: scroll; overflow-y: auto;
    max-height: 560px; max-width: 100%;
    scrollbar-gutter: stable;
  }
  .table-view table {
    border-collapse: collapse; font-size: 12px;
    width: max-content; min-width: 100%;
    background: white;
  }
  /* 让父容器不阻止横向滚动 */
  .card > .table-view, .card #tableView { min-width: 0; }
  /* macOS 默认隐藏滚动条 → 强制始终显示 */
  .table-view::-webkit-scrollbar {
    width: 10px; height: 10px;
    -webkit-appearance: none;
  }
  .table-view::-webkit-scrollbar-track {
    background: #f1f5f9; border-radius: 5px;
  }
  .table-view::-webkit-scrollbar-thumb {
    background: #cbd5e1; border-radius: 5px;
    border: 2px solid #f1f5f9;
  }
  .table-view::-webkit-scrollbar-thumb:hover { background: #94a3b8; }
  .table-view::-webkit-scrollbar-corner { background: #f1f5f9; }
  /* Firefox */
  .table-view { scrollbar-width: thin; scrollbar-color: #cbd5e1 #f1f5f9; }
  .table-view thead th {
    position: sticky; top: 0;
    background: #f8fafc; color: var(--text);
    text-align: left; padding: 8px 12px;
    border-bottom: 2px solid var(--border-strong);
    font-weight: 600; white-space: nowrap;
    font-family: ui-monospace, "SF Mono", Menlo, monospace;
    font-size: 11px;
  }
  .table-view thead .th-type {
    font-family: -apple-system, "PingFang SC", sans-serif;
    font-size: 10px; color: var(--text-3);
    font-weight: normal; margin-top: 2px;
    background: #eef2ff; color: #6366f1;
    display: inline-block; padding: 1px 6px; border-radius: 3px;
  }
  .table-view tbody td {
    padding: 7px 12px; border-bottom: 1px solid #f1f5f9;
    vertical-align: top; white-space: nowrap;
    max-width: 320px; overflow: hidden; text-overflow: ellipsis;
  }
  .table-view tbody td.id-col {
    font-family: ui-monospace, "SF Mono", Menlo, monospace;
    color: var(--text-3); font-size: 11px;
  }
  .table-view tbody tr:hover td { background: #f8fafc; }
  .table-view tbody tr:nth-child(even) td { background: #fafafa; }
  .table-view tbody tr:nth-child(even):hover td { background: #f1f5f9; }
  .table-view tbody tr[data-row-idx] {
    cursor: pointer; user-select: none;
  }
  .table-view tbody tr.row-selected td,
  .table-view tbody tr.row-selected:nth-child(even) td,
  .table-view tbody tr.row-selected:hover td {
    background: var(--primary-soft) !important;
  }
  .table-view tbody tr.row-selected td.id-col {
    color: var(--primary) !important; font-weight: 500;
  }
  .selection-toolbar {
    display: flex; justify-content: space-between; align-items: center;
    padding: 8px 12px; background: var(--primary-soft);
    border-bottom: 1px solid var(--primary-border);
    font-size: 12px; color: var(--primary); font-weight: 500;
    position: sticky; top: 0; z-index: 2;
    display: none;
  }
  .selection-toolbar.active { display: flex; }
  .selection-toolbar button {
    background: var(--primary); color: white;
    padding: 4px 12px; font-size: 12px; border-radius: 4px;
    border: none; cursor: pointer;
  }
  .selection-toolbar button:hover { background: var(--primary-hover); }
  .selection-toolbar button.secondary {
    background: white; color: var(--primary);
    border: 1px solid var(--primary-border);
  }
  .selection-toolbar button.secondary:hover { background: #f5f7fa; }
  .selection-toolbar .btn-group { display: flex; gap: 6px; }
  .table-meta {
    padding: 6px 12px; background: #f8fafc;
    border-bottom: 1px solid var(--border);
    font-size: 11px; color: var(--text-3);
    position: sticky; top: 0; z-index: 1;
  }

  /* ---------- 从 Excel/CSV 导入 小部件 ---------- */
  .import-section {
    margin-top: 12px;
    border: 1px dashed var(--primary-border);
    border-radius: var(--radius-sm);
  }
  .import-toggle {
    width: 100%; text-align: left; background: var(--primary-soft);
    color: var(--primary); border: none; padding: 10px 14px;
    font-size: 13px; cursor: pointer; font-weight: 500;
    border-radius: var(--radius-sm);
  }
  .import-toggle:hover { background: #e0e7ff; }
  .import-panel { padding: 14px; }
  .import-config-row {
    display: grid; grid-template-columns: 110px 1fr; gap: 10px;
    align-items: center; margin-bottom: 10px;
  }
  .import-config-row > label {
    font-size: 12px; color: var(--text-2); text-align: right;
  }
  .import-mapping-table {
    width: 100%; border-collapse: collapse; font-size: 12px;
    margin-top: 6px; background: white; border: 1px solid var(--border);
    border-radius: var(--radius-sm); overflow: hidden;
  }
  .import-mapping-table th {
    background: #f8fafc; padding: 6px 10px; text-align: left;
    color: var(--text-2); font-weight: 500; font-size: 11px;
    border-bottom: 1px solid var(--border);
  }
  .import-mapping-table td {
    padding: 6px 10px; border-bottom: 1px solid #f1f5f9;
    vertical-align: middle;
  }
  .import-mapping-table tr:last-child td { border-bottom: none; }
  .import-mapping-table .arrow { color: var(--text-3); text-align: center; }
  .import-mapping-table select { width: 100%; padding: 4px 8px; font-size: 12px; }
  .import-mapping-table .col-name {
    font-family: ui-monospace, "SF Mono", Menlo, monospace; color: var(--text);
  }
  .import-mapping-table .preview-val {
    color: var(--text-3); font-size: 11px; padding-left: 8px;
  }
  .import-preview {
    margin-top: 10px;
    max-height: 240px; max-width: 100%;
    overflow-y: auto; overflow-x: auto;
    border: 1px solid var(--border); border-radius: var(--radius-sm);
    background: white; min-width: 0;
  }
  .import-preview table {
    width: max-content; min-width: 100%;
    border-collapse: collapse; font-size: 12px;
  }
  /* 防止 grid 子元素撑爆容器 */
  .import-config-row > div { min-width: 0; }
  .import-preview th, .import-preview td {
    padding: 5px 10px; border-bottom: 1px solid #f1f5f9;
    text-align: left; white-space: nowrap;
  }
  .import-preview th {
    background: #f8fafc; position: sticky; top: 0;
    font-size: 11px; color: var(--text-2); font-weight: 500;
  }
  .import-progress {
    margin-top: 10px; padding: 10px;
    background: #1e293b; color: #e2e8f0;
    border-radius: var(--radius-sm);
    font-family: ui-monospace, "SF Mono", Menlo, monospace;
    font-size: 11px; line-height: 1.6; max-height: 240px; overflow: auto;
    display: none;
  }
  .import-progress.active { display: block; }
  .import-progress .ok { color: #4ec9b0; }
  .import-progress .err { color: #f48771; }
  .import-section input[type="file"] {
    padding: 6px; font-size: 12px;
  }

  /* ---------- 按内容搜索 recordId 小部件 ---------- */
  .filter-search { margin-top: 8px; }
  .filter-search-toggle {
    color: var(--primary); background: none; border: none;
    font-size: 12px; padding: 4px 0; cursor: pointer;
    text-decoration: underline dotted; text-underline-offset: 3px;
  }
  .filter-search-toggle:hover { color: var(--primary-hover); background: none; }
  .filter-search-panel {
    margin-top: 8px; padding: 14px;
    background: #fafbfc; border: 1px solid var(--border);
    border-radius: var(--radius-sm);
  }
  .filter-search-panel .row {
    display: flex; gap: 8px; align-items: center; margin-bottom: 8px;
    flex-wrap: wrap;
  }
  .filter-search-panel .row label {
    font-size: 12px; color: var(--text-2); min-width: 60px;
  }
  .filter-search-panel input { flex: 1; min-width: 200px; }
  .filter-search-panel .examples {
    font-size: 11px; color: var(--text-3); margin-top: 4px;
    line-height: 1.7;
  }
  .filter-search-panel .examples code {
    background: white; padding: 1px 5px; border-radius: 3px;
    border: 1px solid var(--border); font-size: 11px;
  }
  .filter-mode-tabs {
    display: flex; gap: 3px; background: white;
    padding: 3px; border-radius: var(--radius-sm); width: fit-content;
    margin-bottom: 12px; border: 1px solid var(--border);
  }
  .filter-mode-tabs button {
    background: transparent; color: var(--text-2);
    padding: 5px 14px; border-radius: 4px; font-size: 12px;
    border: none; cursor: pointer;
  }
  .filter-mode-tabs button.active {
    background: var(--primary-soft); color: var(--primary);
  }
  .filter-cond-row {
    display: grid;
    grid-template-columns: 1.5fr 1fr 1.5fr 32px;
    gap: 6px; margin-bottom: 6px; align-items: center;
  }
  .filter-cond-row select, .filter-cond-row input {
    padding: 5px 8px; font-size: 12px;
  }
  .filter-cond-row .del-btn {
    background: transparent; color: var(--text-3);
    border: none; padding: 4px; font-size: 14px;
    cursor: pointer; border-radius: 4px;
  }
  .filter-cond-row .del-btn:hover { background: var(--danger-soft); color: var(--danger); }
  .filter-cond-row .empty-value {
    color: var(--text-3); font-size: 12px; padding: 5px 8px;
    background: #f5f7fa; border-radius: 4px; text-align: center;
  }
  .filter-connector {
    margin-bottom: 8px; font-size: 12px; color: var(--text-2);
  }
  .filter-connector select {
    padding: 3px 8px; font-size: 12px; min-width: 0; width: auto;
    border-radius: 4px;
  }
  .filter-formula-preview {
    margin-top: 10px; padding: 8px 12px;
    background: #1e293b; color: #e2e8f0;
    border-radius: var(--radius-sm);
    font-family: ui-monospace, "SF Mono", Menlo, monospace;
    font-size: 11px; word-break: break-all;
  }
  .filter-formula-preview .label {
    color: #94a3b8; margin-right: 6px;
  }
  .filter-search-results {
    margin-top: 10px; padding: 10px;
    background: white; border: 1px solid var(--border);
    border-radius: var(--radius-sm); font-size: 12px;
    display: none;
  }
  .filter-search-results.active { display: block; }
  .filter-search-results ul {
    margin: 6px 0 8px; padding-left: 20px;
    font-family: ui-monospace, "SF Mono", Menlo, monospace;
    color: var(--text-3); font-size: 11px;
    max-height: 140px; overflow-y: auto;
  }
  .filter-search-results .actions { display: flex; gap: 6px; margin-top: 8px; }
  .filter-search-results.error {
    background: #fee2e2; border-color: #fecaca; color: var(--danger);
  }

  /* ---------- 主操作行（发送 + 查全表） ---------- */
  .primary-actions {
    display: flex; align-items: center; gap: 10px;
    margin-bottom: 12px; flex-wrap: wrap;
  }
  .send-secondary {
    background: white; color: var(--primary);
    border: 1px solid var(--primary-border);
    padding: 10px 20px; font-size: 14px;
    border-radius: var(--radius-sm); cursor: pointer;
    transition: all 0.15s; font-family: inherit;
  }
  .send-secondary:hover {
    background: var(--primary-soft);
  }
  .hint-inline {
    color: var(--text-3); font-size: 12px; margin-left: 4px;
  }

  /* ---------- 进度条 ---------- */
  .progress-bar {
    height: 3px; border-radius: 2px;
    background: var(--border); margin-bottom: 12px;
    position: relative; overflow: hidden;
    display: none;
  }
  .progress-bar.active { display: block; }
  .progress-bar::before {
    content: ""; position: absolute; inset: 0;
    background: linear-gradient(90deg, transparent, var(--primary), transparent);
    animation: progress-slide 1.2s linear infinite;
  }
  @keyframes progress-slide {
    0% { transform: translateX(-100%); }
    100% { transform: translateX(100%); }
  }
  button[disabled], button.loading {
    opacity: 0.55; cursor: wait !important;
  }
  button.loading { position: relative; }

  /* ---------- 空状态 ---------- */
  .empty-state {
    color: var(--text-3); font-size: 13px;
    padding: 20px; text-align: center;
    border: 1px dashed var(--border-strong); border-radius: var(--radius-sm);
  }

  /* ---------- 定时任务表 ---------- */
  .schedules-table {
    width: 100%; border-collapse: collapse; font-size: 13px;
  }
  .schedules-table th {
    text-align: left; padding: 8px 10px; font-size: 11px;
    color: var(--text-3); font-weight: 500; text-transform: uppercase;
    border-bottom: 1px solid var(--border); letter-spacing: 0.3px;
  }
  .schedules-table td {
    padding: 10px; border-bottom: 1px solid #f1f5f9;
    vertical-align: middle;
  }
  .schedules-table tr:last-child td { border-bottom: none; }
  .schedules-table .name-col { font-weight: 500; }
  .schedules-table .status-ok { color: var(--success); }
  .schedules-table .status-err { color: var(--danger); }
  .schedules-table .status-pending { color: var(--text-3); }
  .schedules-table .row-actions {
    display: flex; gap: 4px; flex-wrap: wrap;
  }
  .switch {
    position: relative; display: inline-block;
    width: 36px; height: 20px;
  }
  .switch input { opacity: 0; width: 0; height: 0; }
  .switch .slider {
    position: absolute; cursor: pointer; inset: 0;
    background: #cbd5e1; border-radius: 20px;
    transition: 0.2s;
  }
  .switch .slider:before {
    position: absolute; content: ""; height: 14px; width: 14px;
    left: 3px; bottom: 3px; background: white;
    border-radius: 50%; transition: 0.2s;
  }
  .switch input:checked + .slider { background: var(--primary); }
  .switch input:checked + .slider:before { transform: translateX(16px); }
</style>
</head>
<body>
<div class="container">
  <div class="top-bar">
    <h1><span class="icon">A</span>AITable API 调试工具</h1>
    <a class="doc-link" href="https://developers.aitable.ai/api/cn/reference/" target="_blank">
      📖 官方 API 文档
    </a>
  </div>

  <!-- 通用配置 -->
  <div class="card">
    <div class="card-head">
      <h2>通用配置</h2>
      <span class="meta">跨所有操作共用</span>
    </div>
    <div class="grid-2">
      <div class="field">
        <div class="field-label">接口地址 <span class="api-name">baseUrl</span> <span class="req">*</span></div>
        <input type="text" id="baseUrl" list="baseUrlPresets"
               placeholder="https://你的域名/fusion/v1（自部署版填公司域名）"
               value="https://apitable.yottastudios.com/fusion/v1">
        <datalist id="baseUrlPresets">
          <option value="https://apitable.yottastudios.com/fusion/v1">Yotta 自部署</option>
          <option value="https://aitable.ai/fusion/v1">国际版（公网）</option>
          <option value="https://vika.cn/fusion/v1">国内版（公网）</option>
        </datalist>
        <div style="font-size: 11px; color: var(--text-3); margin-top: 4px;">
          自部署版规则：把表格 URL 中 <code style="background:#f3f4f6;padding:1px 4px;border-radius:3px;">/workbench/...</code> 之前那段加 <code style="background:#f3f4f6;padding:1px 4px;border-radius:3px;">/fusion/v1</code>
        </div>
      </div>
      <div class="field">
        <div class="field-label">API 密钥 <span class="api-name">token</span> <span class="req">*</span></div>
        <input type="password" id="token" placeholder="在 AITable 用户中心 → 开发者配置 申请">
      </div>
    </div>
    <div class="actions" style="margin-top: 12px;">
      <button class="ghost small" onclick="saveBaseConfig()">💾 保存</button>
      <button class="ghost small" onclick="loadBaseConfig()">📂 加载</button>
    </div>
  </div>

  <!-- 操作 + 模板 -->
  <div class="card">
    <div class="card-head">
      <h2>操作 & 模板</h2>
    </div>
    <div class="grid-2">
      <div class="field">
        <div class="field-label">操作类型</div>
        <select id="operation">
          <option value="get_records">查询记录（GET records）</option>
          <option value="create">创建记录（CREATE）</option>
          <option value="update">更新记录（UPDATE）</option>
          <option value="delete">删除记录（DELETE）</option>
          <option value="get_fields">查询字段列表（GET fields）</option>
          <option value="get_views">查询视图列表（GET views）</option>
          <option value="get_spaces">查询空间站列表（GET spaces）</option>
        </select>
      </div>
      <div class="field">
        <div class="field-label">模板（保存表格 ID + 操作 + 参数）</div>
        <div class="template-bar">
          <select id="templateSelect">
            <option value="">— 选择已保存的模板 —</option>
          </select>
          <button class="small" onclick="loadTemplate()">加载</button>
        </div>
      </div>
    </div>
    <div class="actions" style="margin-top: 12px;">
      <button class="ghost small" onclick="saveAsTemplate()">＋ 另存为模板</button>
      <button class="ghost small" onclick="overwriteTemplate()">↻ 覆盖当前模板</button>
      <button class="ghost small" onclick="deleteTemplate()" style="color: var(--danger);">✕ 删除模板</button>
    </div>
  </div>

  <!-- 参数 -->
  <div class="card">
    <div class="card-head">
      <h2>参数</h2>
      <span class="meta" id="paramsMeta"></span>
    </div>
    <div id="params"></div>
  </div>

  <!-- 定时任务 -->
  <div class="card">
    <div class="card-head">
      <h2>🔁 定时任务</h2>
      <span class="meta">仅在工具运行时执行</span>
    </div>
    <div id="schedulesList"></div>
    <div class="actions" style="margin-top: 12px;">
      <button class="secondary small" onclick="addScheduleFromCurrent()">＋ 把当前操作设为定时任务</button>
      <button class="ghost small" onclick="refreshSchedules()">🔄 刷新</button>
      <button class="ghost small" onclick="showScheduleLogs()">📜 查看执行日志</button>
    </div>
  </div>

  <!-- 执行 -->
  <div class="card">
    <div class="card-head">
      <h2>执行 & 响应</h2>
      <div class="actions">
        <button class="ghost small" onclick="document.getElementById('result').textContent=''">清空</button>
      </div>
    </div>
    <div class="primary-actions">
      <button class="send" id="btnSend" onclick="sendRequest()">🚀 发送请求</button>
      <button class="send-secondary" id="btnQueryAll" onclick="queryAllRecords()">📋 查询整张表</button>
      <span class="hint-inline">查询整张表可以一键拿到所有记录的 recordId</span>
    </div>
    <div id="progressBar" class="progress-bar"></div>
    <div id="responseToolbar" class="response-toolbar" style="display:none;">
      <div class="view-toggle">
        <button class="active" data-view="json" onclick="switchView('json')">原始 JSON</button>
        <button data-view="table" onclick="switchView('table')">📊 表格视图</button>
      </div>
      <div class="actions">
        <button id="btnRefetch" class="secondary small" onclick="refetchFullRecords()" style="display:none;">🔄 重新查询本次记录</button>
        <button id="btnOpenAitable" class="secondary small" onclick="openInAitable()" style="display:none;">🔗 在 APITable 中打开</button>
        <button class="secondary small" onclick="downloadCSV()">⬇ 下载 CSV (Excel)</button>
        <button class="secondary small" onclick="downloadJSON()">⬇ 下载 JSON</button>
      </div>
    </div>
    <div id="result" class="result-box"></div>
    <div id="tableView" class="table-view" style="display:none;"></div>
  </div>
</div>

<script>
// ---------- 字段元数据 ----------
// fullWidth: 该字段独占一行；其他默认按双栏栅格布局
var FORMS = {
  "get_records": [
    { key: "datasheetId", label: "表格 ID", apiName: "datasheetId", required: true, fullWidth: true,
      placeholder: "dst 开头，从表格页面 URL 中复制",
      help: "表格的唯一标识。从 AITable 中打开任一表格，浏览器地址栏 /workbench/dst.../viw... 中 dst 开头的那段就是 datasheetId。" },
    { key: "viewId", label: "视图 ID", apiName: "viewId",
      placeholder: "可选；viw 开头",
      help: "视图 ID。指定视图则返回的 fields 顺序和视图保持一致，隐藏的字段不会返回。地址栏 viw 开头那段。" },
    { key: "fieldKey", label: "字段标识方式", apiName: "fieldKey", "default": "name", type: "select", options: ["name", "id"],
      help: "查询字段和返回字段时所用的 key。默认 name（按列名），指定为 id 时将以 fieldId 作为查询和返回方式。" },
    { key: "pageSize", label: "每页数量", apiName: "pageSize",
      placeholder: "1-1000，默认 100",
      help: "每页返回多少条记录。默认每页返回 100 条记录。取值范围为 1~1000 的整数。" },
    { key: "pageNum", label: "页码", apiName: "pageNum",
      placeholder: "从 1 开始",
      help: "指定分页的页码，与参数 pageSize 配合使用。例如 pageSize=1000&pageNum=2，返回 1001~2000 之间的记录。" },
    { key: "maxRecords", label: "总条数上限", apiName: "maxRecords",
      placeholder: "可选",
      help: "总共返回多少条记录。如果 maxRecords 与 pageSize 同时使用，且 maxRecords 的值小于总记录数，则只生效 maxRecords 的设置。" },
    { key: "recordIds", label: "指定记录 ID", apiName: "recordIds",
      placeholder: "rec1,rec2,...",
      help: "返回一个或多个指定的记录。多个记录 ID 用逗号分隔。返回结果按照传入 recordId 的顺序排序，无分页，每次最多返回 1000 条。" },
    { key: "fields", label: "只返回这些字段", apiName: "fields",
      placeholder: "字段名1,字段名2",
      help: "限制在返回的记录结果只包含指定的字段。多个字段名用逗号分隔。" },
    { key: "filterByFormula", label: "筛选公式", apiName: "filterByFormula", fullWidth: true,
      placeholder: '例：{状态}="完成"',
      help: '使用智能公式来筛选记录。常用例子：{状态}="完成"、{重要度}>=4、AND({状态}="进行中",{负责人}="张三")。如果与 viewId 同时使用，则返回指定视图中满足公式的所有记录。' }
  ],
  "create": [
    { key: "datasheetId", label: "表格 ID", apiName: "datasheetId", required: true, fullWidth: true,
      placeholder: "dst 开头",
      help: "要创建记录的目标表格的 ID。" },
    { key: "viewId", label: "视图 ID（可选）", apiName: "viewId",
      placeholder: "viw 开头；不填则按默认视图",
      help: "可选。指定视图后，新记录会按视图的字段范围和顺序插入。注意：这不会影响响应——响应永远只返回刚创建的记录。" },
    { key: "fieldKey", label: "字段标识方式", apiName: "fieldKey", "default": "name", type: "select", options: ["name", "id"],
      help: "字段映射方式。name = 用列名作为 key（推荐，直观），id = 用 fieldId 作为 key（避免列改名后报错）。" },
    { key: "records", label: "记录数据", type: "records", mode: "create", fullWidth: true, showFileImport: true,
      help: "需要创建的记录数据。\n\n💡 三种填法：\n  1) 手动加：每点一次「+ 添加一条记录」加一条\n  2) 从 Excel/CSV 导入：点下方「📊 从 Excel/CSV 批量导入」\n  3) 直接写 JSON：切到「JSON 模式」\n\n注：API 单次请求最多 10 条；导入工具会自动分批发送，无 10 条限制。" }
  ],
  "update": [
    { key: "datasheetId", label: "表格 ID", apiName: "datasheetId", required: true, fullWidth: true,
      placeholder: "dst 开头",
      help: "要更新记录的目标表格的 ID。" },
    { key: "viewId", label: "视图 ID（可选）", apiName: "viewId",
      placeholder: "viw 开头；不填则按默认视图",
      help: "可选。指定视图后，操作的字段范围限制在该视图。不影响响应——响应永远只返回刚更新的记录。" },
    { key: "fieldKey", label: "字段标识方式", apiName: "fieldKey", "default": "name", type: "select", options: ["name", "id"],
      help: "字段映射方式。name = 用列名（推荐），id = 用 fieldId。" },
    { key: "records", label: "要更新的记录", type: "records", mode: "update", fullWidth: true, showFilterSearch: true,
      help: "需要更新的记录列表。每条必须含 recordId（rec 开头），加上要修改的字段。值设为 null 可清空字段。最多 10 条。\n\n💡 三种填法：\n  1) 手动加每条：点「+ 添加一条记录」→ 填 recordId 和字段\n  2) 批量修改：先在「第 1 条」里填好要修改的字段（作为模板）→ 点下方「🔍 按内容搜索」→ 找到匹配记录后点「批量生成」→ 用同一组字段更新所有匹配记录\n  3) 从查询整张表选行：上方「📋 查询整张表」→ 表格视图选行 → 「↩ 用选中 ID 填到 update/delete」" }
  ],
  "delete": [
    { key: "datasheetId", label: "表格 ID", apiName: "datasheetId", required: true, fullWidth: true,
      placeholder: "dst 开头",
      help: "要删除记录所在的表格 ID。" },
    { key: "viewId", label: "视图 ID（可选）", apiName: "viewId",
      placeholder: "viw 开头；不影响删除本身",
      help: "可选。AITable 的 DELETE 接口本身只接受 recordIds（不接 viewId），这里填只是为了方便：删除后点「📋 查询整张表」时会用这个 viewId 来限定显示范围。" },
    { key: "recordIds", label: "要删除的记录 ID", apiName: "recordIds", required: true, fullWidth: true,
      placeholder: "rec1,rec2,...（最多 10 个）",
      showFilterSearch: true,
      help: "要删除的记录 ID 列表，多个用英文逗号分隔，例如：rec4zxfWB5uyM,reclNflLgtzjY。一次最多 10 个。\n\n💡 三种填法：\n  1) 点上方「📋 查询整张表」→ 表格视图选行 → 「↩ 用选中 ID 填到 update/delete」\n  2) 点下方「🔍 按内容搜索」→ 写筛选公式 → 自动填入匹配的 recordId\n  3) 直接手动粘贴 recordId（rec 开头那串）\n\n注：SDK 里的 .filter(标题=xxx).delete() 是 SDK 底层先查后删，REST API 本身只接受 recordIds。" }
  ],
  "get_fields": [
    { key: "datasheetId", label: "表格 ID", apiName: "datasheetId", required: true,
      placeholder: "dst 开头",
      help: "目标表格 ID。返回该表格的所有字段（列）的元数据：字段 ID、名称、类型、配置等。" },
    { key: "viewId", label: "视图 ID", apiName: "viewId",
      placeholder: "可选",
      help: "视图 ID。指定后返回的字段顺序和该视图一致，被视图隐藏的字段不会返回。" }
  ],
  "get_views": [
    { key: "datasheetId", label: "表格 ID", apiName: "datasheetId", required: true, fullWidth: true,
      placeholder: "dst 开头",
      help: "目标表格 ID。返回该表格的所有视图（含视图 ID、名称、类型）。" }
  ],
  "get_spaces": []
};

// 操作类型描述（显示在参数区右上角）
var OP_META = {
  get_records: "GET /datasheets/{id}/records",
  create: "POST /datasheets/{id}/records",
  update: "PATCH /datasheets/{id}/records",
  delete: "DELETE /datasheets/{id}/records",
  get_fields: "GET /datasheets/{id}/fields",
  get_views: "GET /datasheets/{id}/views",
  get_spaces: "GET /spaces"
};

// ---------- 记录编辑器状态 ----------
var recordsState = { create: null, update: null };
var jsonModeState = { create: false, update: false };

// ---------- 跨操作切换时保留的同名参数 ----------
// 用户切换操作类型时，相同字段名的值会自动继承（如 datasheetId, viewId, fieldKey）
// 但如果 datasheetId 被改了，跟表格挂钩的参数（viewId/recordIds/fields/filterByFormula）会清掉
var sharedParams = {};
var lastSharedDatasheetId = "";

function captureSharedParams() {
  var op = document.getElementById("operation").value;
  var fields = FORMS[op] || [];
  fields.forEach(function(f) {
    if (f.type === "records") return;
    var el = document.getElementById("param_" + f.key);
    if (!el) return;
    sharedParams[f.key] = el.value;
  });
  var currentDst = sharedParams.datasheetId || "";
  if (lastSharedDatasheetId && currentDst !== lastSharedDatasheetId) {
    delete sharedParams.viewId;
    delete sharedParams.recordIds;
    delete sharedParams.fields;
    delete sharedParams.filterByFormula;
  }
  lastSharedDatasheetId = currentDst;
}

function restoreSharedParams() {
  var op = document.getElementById("operation").value;
  var fields = FORMS[op] || [];
  fields.forEach(function(f) {
    if (f.type === "records") return;
    if (!(f.key in sharedParams)) return;
    var v = sharedParams[f.key];
    if (v === "" || v == null) return;
    var el = document.getElementById("param_" + f.key);
    if (el) el.value = v;
  });
}

function defaultRecord(mode) {
  var r = { fields: [{ name: "", value: "" }] };
  if (mode === "update") r.recordId = "";
  return r;
}

function ensureRecordsState(mode) {
  if (!recordsState[mode]) recordsState[mode] = [defaultRecord(mode)];
  return recordsState[mode];
}

function coerceValue(s) {
  if (s === "") return null;
  if (s === "true") return true;
  if (s === "false") return false;
  if (/^-?\d+$/.test(s)) return parseInt(s, 10);
  if (/^-?\d*\.\d+$/.test(s)) return parseFloat(s);
  if ((s.charAt(0) === "[" && s.charAt(s.length - 1) === "]") ||
      (s.charAt(0) === "{" && s.charAt(s.length - 1) === "}")) {
    try { return JSON.parse(s); } catch (e) {}
  }
  return s;
}

// 检查一条记录里有没有重复字段名（JSON 对象 key 唯一，重复会互相覆盖）
function detectDuplicateFields(rec) {
  var nameToIndices = {};
  (rec.fields || []).forEach(function(f, idx) {
    if (!f.name) return;
    if (!nameToIndices[f.name]) nameToIndices[f.name] = [];
    nameToIndices[f.name].push(idx);
  });
  var dupIndices = {};
  var dupNames = [];
  Object.keys(nameToIndices).forEach(function(name) {
    if (nameToIndices[name].length > 1) {
      dupNames.push(name + " ×" + nameToIndices[name].length);
      nameToIndices[name].forEach(function(i) { dupIndices[i] = true; });
    }
  });
  return { indices: dupIndices, names: dupNames };
}

function buildRecordsPayload(mode) {
  var records = recordsState[mode] || [];
  return records.map(function(rec) {
    var fieldsObj = {};
    rec.fields.forEach(function(f) {
      if (!f.name) return;
      var v = f.value;
      var t = f.type || "自动";
      if (t === "文本") fieldsObj[f.name] = v;
      else if (t === "数字") fieldsObj[f.name] = v === "" ? null : Number(v);
      else if (t === "布尔") fieldsObj[f.name] = v === "true" || v === "1" || v.toLowerCase() === "true";
      else if (t === "JSON") {
        try { fieldsObj[f.name] = JSON.parse(v); } catch (e) { fieldsObj[f.name] = v; }
      } else fieldsObj[f.name] = coerceValue(v);
    });
    var out = { fields: fieldsObj };
    if (mode === "update") out.recordId = rec.recordId || "";
    return out;
  });
}

function renderRecordsEditor(container, mode) {
  container.innerHTML = "";

  var toolbar = document.createElement("div");
  toolbar.className = "records-toolbar";
  var btnVisual = document.createElement("button");
  btnVisual.textContent = "可视化";
  btnVisual.className = jsonModeState[mode] ? "" : "active";
  btnVisual.onclick = function() { jsonModeState[mode] = false; renderRecordsEditor(container, mode); };
  var btnJson = document.createElement("button");
  btnJson.textContent = "JSON 模式";
  btnJson.className = jsonModeState[mode] ? "active" : "";
  btnJson.onclick = function() { jsonModeState[mode] = true; renderRecordsEditor(container, mode); };
  toolbar.appendChild(btnVisual);
  toolbar.appendChild(btnJson);
  container.appendChild(toolbar);

  if (jsonModeState[mode]) {
    var ta = document.createElement("textarea");
    ta.id = "param_records_json_" + mode;
    ta.rows = 14;
    ta.value = JSON.stringify(buildRecordsPayload(mode), null, 2);
    container.appendChild(ta);
    return;
  }

  var records = ensureRecordsState(mode);

  records.forEach(function(rec, idx) {
    var card = document.createElement("div");
    card.className = "record-card";

    var header = document.createElement("div");
    header.className = "record-card-header";
    var title = document.createElement("span");
    title.className = "record-title";
    title.textContent = "第 " + (idx + 1) + " 条";
    header.appendChild(title);

    if (records.length > 1) {
      var rmRec = document.createElement("button");
      rmRec.className = "danger small";
      rmRec.textContent = "删除该条";
      rmRec.onclick = function() {
        records.splice(idx, 1);
        renderRecordsEditor(container, mode);
      };
      header.appendChild(rmRec);
    }
    card.appendChild(header);

    if (mode === "update") {
      var rid = document.createElement("div");
      rid.className = "recordid-row";
      var rl = document.createElement("label");
      rl.textContent = "记录 ID";
      var ri = document.createElement("input");
      ri.type = "text";
      ri.placeholder = "rec 开头";
      ri.value = rec.recordId || "";
      ri.oninput = function() { rec.recordId = ri.value; };
      rid.appendChild(rl);
      rid.appendChild(ri);
      card.appendChild(rid);
    }

    // 检测重复字段名 → 警告 banner
    var dup = detectDuplicateFields(rec);
    if (dup.names.length > 0) {
      var warn = document.createElement("div");
      warn.className = "duplicate-warning";
      warn.innerHTML = '⚠ <b>这条记录里字段名重复了：' + escapeHtml(dup.names.join("，")) + '</b><br>' +
        '同名字段在 JSON 里会互相覆盖（只保留最后一个值）。<br>' +
        '👉 想创建多条记录请用下方「<b>+ 添加一条记录</b>」按钮，而不是「+ 添加字段」。';
      card.appendChild(warn);
    }

    var head = document.createElement("div");
    head.className = "field-row-head";
    head.innerHTML = "<div>字段名 / 列名</div><div>字段值</div><div>类型</div><div></div>";
    card.appendChild(head);

    rec.fields.forEach(function(field, fi) {
      var row = document.createElement("div");
      row.className = "field-row" + (dup.indices[fi] ? " duplicate" : "");

      var nameI = document.createElement("input");
      nameI.type = "text"; nameI.placeholder = "如：标题";
      nameI.value = field.name;
      nameI.oninput = function() { field.name = nameI.value; };
      // 失焦或回车时重新渲染（这样重复警告能及时显示，又不会打断打字）
      nameI.onchange = function() { renderRecordsEditor(container, mode); };

      var valI = document.createElement("input");
      valI.type = "text"; valI.placeholder = "如：测试内容";
      valI.value = field.value;
      valI.oninput = function() { field.value = valI.value; };

      var typeS = document.createElement("select");
      ["自动", "文本", "数字", "布尔", "JSON"].forEach(function(t) {
        var o = document.createElement("option");
        o.value = t; o.textContent = t; typeS.appendChild(o);
      });
      typeS.value = field.type || "自动";
      typeS.onchange = function() { field.type = typeS.value; };

      var del = document.createElement("button");
      del.className = "del-btn"; del.textContent = "✕";
      del.title = "删除该字段";
      del.onclick = function() {
        rec.fields.splice(fi, 1);
        if (rec.fields.length === 0) rec.fields.push({ name: "", value: "" });
        renderRecordsEditor(container, mode);
      };

      row.appendChild(nameI);
      row.appendChild(valI);
      row.appendChild(typeS);
      row.appendChild(del);
      card.appendChild(row);
    });

    var addField = document.createElement("button");
    addField.className = "ghost small";
    addField.textContent = "＋ 添加字段";
    addField.style.marginTop = "8px";
    addField.onclick = function() {
      rec.fields.push({ name: "", value: "" });
      renderRecordsEditor(container, mode);
    };
    card.appendChild(addField);

    container.appendChild(card);
  });

  if (records.length < 10) {
    var addRec = document.createElement("button");
    addRec.className = "secondary small";
    addRec.textContent = "＋ 添加一条记录";
    addRec.onclick = function() {
      records.push(defaultRecord(mode));
      renderRecordsEditor(container, mode);
    };
    container.appendChild(addRec);
  }
}

// ---------- 渲染参数区 ----------
function renderParams() {
  try {
    var op = document.getElementById("operation").value;
    var fields = FORMS[op] || [];
    var container = document.getElementById("params");
    var meta = document.getElementById("paramsMeta");
    container.innerHTML = "";
    meta.textContent = OP_META[op] || "";

    if (fields.length === 0) {
      var empty = document.createElement("div");
      empty.className = "empty-state";
      empty.textContent = "此操作无需额外参数 — 直接点「发送请求」";
      container.appendChild(empty);
      return;
    }

    var grid = document.createElement("div");
    grid.className = "grid-2";

    for (var i = 0; i < fields.length; i++) {
      var f = fields[i];

      var fieldDiv = document.createElement("div");
      fieldDiv.className = "field" + (f.fullWidth ? " full" : "");

      // Label
      var label = document.createElement("div");
      label.className = "field-label";
      var labelText = document.createElement("span");
      labelText.textContent = f.label;
      label.appendChild(labelText);
      if (f.apiName) {
        var apiName = document.createElement("span");
        apiName.className = "api-name";
        apiName.textContent = f.apiName;
        label.appendChild(apiName);
      }
      if (f.required) {
        var req = document.createElement("span");
        req.className = "req"; req.textContent = "*";
        label.appendChild(req);
      }
      if (f.help) {
        var helpKey = f.key;
        var helpIcon = document.createElement("span");
        helpIcon.className = "help-icon"; helpIcon.textContent = "?";
        helpIcon.title = "点击查看官方说明";
        helpIcon.onclick = function(k) {
          return function(e) {
            e.stopPropagation();
            var p = document.getElementById("help_" + k);
            if (p) p.classList.toggle("open");
          };
        }(helpKey);
        label.appendChild(helpIcon);
      }
      fieldDiv.appendChild(label);

      // Help panel (collapsed)
      if (f.help) {
        var helpP = document.createElement("div");
        helpP.className = "help-panel";
        helpP.id = "help_" + f.key;
        helpP.textContent = f.help;
        fieldDiv.appendChild(helpP);
      }

      // Input
      if (f.type === "records") {
        var ed = document.createElement("div");
        ed.id = "records_editor_" + f.mode;
        fieldDiv.appendChild(ed);
        // 「按内容搜索」小部件（update 模式启用，可批量生成更新记录）
        if (f.showFilterSearch) {
          fieldDiv.appendChild(buildFilterSearchWidget());
        }
        // 「Excel/CSV 批量导入」小部件（create 模式启用）
        if (f.showFileImport) {
          fieldDiv.appendChild(buildFileImportWidget());
        }
        grid.appendChild(fieldDiv);
        renderRecordsEditor(ed, f.mode);
        continue;
      }

      var input;
      if (f.type === "select") {
        input = document.createElement("select");
        for (var j = 0; j < f.options.length; j++) {
          var opt = document.createElement("option");
          opt.value = f.options[j]; opt.textContent = f.options[j];
          input.appendChild(opt);
        }
        if (f["default"]) input.value = f["default"];
      } else {
        input = document.createElement("input");
        input.type = "text";
        if (f["default"]) input.value = f["default"];
      }
      input.id = "param_" + f.key;
      if (f.placeholder) input.placeholder = f.placeholder;
      fieldDiv.appendChild(input);

      // 「按内容搜索 recordId」小部件（仅 delete 的 recordIds 字段启用）
      if (f.showFilterSearch) {
        fieldDiv.appendChild(buildFilterSearchWidget());
      }

      grid.appendChild(fieldDiv);
    }

    container.appendChild(grid);
  } catch (err) {
    var c = document.getElementById("params");
    if (c) c.textContent = "渲染参数出错: " + err.message;
    if (window.console) console.error(err);
  }
}

function getParamValue(key, opMode) {
  if (key === "records") {
    if (jsonModeState[opMode]) {
      var ta = document.getElementById("param_records_json_" + opMode);
      return ta ? ta.value.trim() : "";
    }
    return JSON.stringify(buildRecordsPayload(opMode));
  }
  var el = document.getElementById("param_" + key);
  return el ? el.value.trim() : "";
}

function showResult(text, isError) {
  var el = document.getElementById("result");
  el.textContent = text;
  el.className = "result-box" + (isError ? " error" : "");
}

// ---------- 加载状态 ----------
var _loadingDots = null;
var _loadingStartedAt = 0;

function showLoading(text) {
  document.getElementById("progressBar").classList.add("active");
  // 禁用所有可能触发请求的按钮
  ["btnSend", "btnRefetch", "btnQueryAll"].forEach(function(id) {
    var b = document.getElementById(id);
    if (b) { b.disabled = true; b.classList.add("loading"); }
  });
  // 在结果区显示带计时器的加载提示
  _loadingStartedAt = Date.now();
  var baseText = text || "请求中";
  var dots = 0;
  var update = function() {
    var elapsed = ((Date.now() - _loadingStartedAt) / 1000).toFixed(1);
    var d = ".".repeat((dots % 4));
    showResult("⏳ " + baseText + d + "  (" + elapsed + "s)", false);
    dots++;
  };
  update();
  if (_loadingDots) clearInterval(_loadingDots);
  _loadingDots = setInterval(update, 400);
}

function hideLoading() {
  document.getElementById("progressBar").classList.remove("active");
  ["btnSend", "btnRefetch", "btnQueryAll"].forEach(function(id) {
    var b = document.getElementById(id);
    if (b) { b.disabled = false; b.classList.remove("loading"); }
  });
  if (_loadingDots) { clearInterval(_loadingDots); _loadingDots = null; }
}

// ---------- 响应数据 + 表格视图 + 导出 ----------
var lastResponse = null;   // 完整响应对象 {status, request, response}
var lastRecords = null;    // 提取出的 records 数组（如果是查询记录响应）
var lastFieldSchema = null;// 表格字段结构 [{id, name, type, ...}]
var lastFieldKey = "name"; // 上次记录里的 fields 是按 name 还是 id 索引
var lastDatasheetId = "";  // 上一次请求用的表格 ID（refetch / 打开外链用）
var lastBaseUrl = "";      // 上一次请求用的 baseUrl
var lastToken = "";        // 上一次请求用的 token

// 把当前可用信息整合成"表格列定义"：每列有 (key, header, type)
function getDisplayColumns() {
  if (lastFieldSchema && lastFieldSchema.length > 0) {
    return lastFieldSchema.map(function(f) {
      var key = (lastFieldKey === "id") ? f.id : f.name;
      return { key: key, header: f.name, type: f.type || null };
    });
  }
  // 兜底：从记录里收集
  var fieldSet = {};
  (lastRecords || []).forEach(function(r) {
    if (r.fields) Object.keys(r.fields).forEach(function(k) { fieldSet[k] = true; });
  });
  return Object.keys(fieldSet).map(function(k) { return { key: k, header: k, type: null }; });
}

// 字段类型 → 中文短名（用在列头）
var FIELD_TYPE_LABELS = {
  Text: "文本", SingleText: "单行文本", LongText: "多行",
  SingleSelect: "单选", MultiSelect: "多选",
  Number: "数字", Currency: "货币", Percent: "百分比", Rating: "评分",
  Checkbox: "勾选",
  DateTime: "日期", CreatedTime: "创建时间", LastModifiedTime: "修改时间",
  Member: "成员", CreatedBy: "创建人", LastModifiedBy: "修改人",
  Attachment: "附件",
  URL: "URL", Email: "邮箱", Phone: "电话",
  AutoNumber: "自增ID",
  MagicLink: "神奇关联", MagicLookUp: "神奇引用",
  Formula: "智能公式", WorkDoc: "工作文档", Cascader: "级联", Button: "按钮"
};

function fieldTypeLabel(type) {
  if (!type) return "";
  return FIELD_TYPE_LABELS[type] || type;
}

// 后台静默拉取字段 schema（在已经显示了响应之后），到了再重新渲染表格
function maybeFetchSchema() {
  if (!lastRecords || lastRecords.length === 0) return;
  if (!lastDatasheetId || !lastToken) return;
  if (lastFieldSchema) return;  // 已经有了

  fetch("/api/call", {
    method: "POST", headers: {"Content-Type": "application/json"},
    body: JSON.stringify({
      op: "get_fields", token: lastToken, baseUrl: lastBaseUrl, datasheetId: lastDatasheetId,
      params: { datasheetId: lastDatasheetId }
    })
  }).then(function(r) { return r.json(); })
    .then(function(data) {
      if (data && data.response && data.response.data
          && Array.isArray(data.response.data.fields)) {
        lastFieldSchema = data.response.data.fields;
        // 如果当前是表格视图，刷新一下
        if (document.getElementById("tableView").style.display !== "none") {
          renderTableView();
        }
      }
    })
    .catch(function() {/* 静默失败，不打扰用户 */});
}

function extractRecords(resp) {
  if (!resp || typeof resp !== "object") return null;
  // 标准位置：response.data.records
  if (resp.data && Array.isArray(resp.data.records)) {
    var r = resp.data.records;
    if (r.length === 0 || (r[0] && typeof r[0] === "object" && "fields" in r[0])) return r;
  }
  // 兜底：response.records 直接在外层
  if (Array.isArray(resp.records)) {
    var r2 = resp.records;
    if (r2.length === 0 || (r2[0] && typeof r2[0] === "object" && "fields" in r2[0])) return r2;
  }
  return null;
}

function formatTimestamp(ms) {
  if (!ms) return "";
  var d = new Date(ms);
  if (isNaN(d.getTime())) return String(ms);
  function pad(n) { return n < 10 ? "0" + n : n; }
  return d.getFullYear() + "-" + pad(d.getMonth()+1) + "-" + pad(d.getDate())
       + " " + pad(d.getHours()) + ":" + pad(d.getMinutes()) + ":" + pad(d.getSeconds());
}

function formatCellValue(v, type) {
  if (v == null || v === "") return "";

  // 按字段类型格式化
  if (type === "DateTime" || type === "CreatedTime" || type === "LastModifiedTime") {
    if (typeof v === "number") return formatTimestamp(v);
    if (typeof v === "string" && /^\d{10,}$/.test(v)) return formatTimestamp(parseInt(v, 10));
  }
  if (type === "Checkbox") {
    return v === true || v === "true" ? "✓" : "";
  }
  if (type === "Percent" && typeof v === "number") {
    return (v * 100).toFixed(2).replace(/\.?0+$/, "") + "%";
  }
  if (type === "Currency" && typeof v === "number") {
    return v.toLocaleString();
  }
  if (type === "Rating" && typeof v === "number") {
    return "★".repeat(v) + "☆".repeat(Math.max(0, 5 - v));
  }
  if ((type === "Member" || type === "CreatedBy" || type === "LastModifiedBy") && Array.isArray(v)) {
    return v.map(function(x) { return x && (x.name || x.userName) || ""; }).filter(Boolean).join(", ");
  }
  if (type === "Attachment" && Array.isArray(v)) {
    return v.map(function(x) {
      if (!x) return "";
      return (x.name || "附件") + (x.size ? " (" + (x.size/1024).toFixed(1) + "KB)" : "");
    }).filter(Boolean).join(", ");
  }

  // 默认逻辑
  if (typeof v === "string") return v;
  if (typeof v === "number" || typeof v === "boolean") return String(v);
  if (Array.isArray(v)) {
    return v.map(function(x) {
      if (x == null) return "";
      if (typeof x === "object") {
        if (x.url) return (x.name || "") + " " + x.url;
        if (x.title) return x.title;
        if (x.name) return x.name;
        return JSON.stringify(x);
      }
      return String(x);
    }).join(", ");
  }
  if (typeof v === "object") return JSON.stringify(v);
  return String(v);
}

function escapeHtml(s) {
  return String(s).replace(/[&<>"']/g, function(c) {
    return { "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;", "'": "&#39;" }[c];
  });
}

// 行选中状态：每次重新渲染清空
var selectedRows = new Set();
var lastClickedRowIdx = -1;
var dragStartRowIdx = -1;
var isDraggingSelection = false;
// 拖动模式：null=尚未决定 / "select"=多选 / "pan"=横向滚动
var _dragMode = null;
var _dragStartX = 0, _dragStartY = 0;
var _initialScrollLeft = 0;
var _DRAG_THRESHOLD = 8;

function renderTableView() {
  var container = document.getElementById("tableView");
  if (!lastRecords || lastRecords.length === 0) {
    container.innerHTML = '<div style="padding:24px;text-align:center;color:var(--text-3);">暂无数据</div>';
    selectedRows.clear();
    return;
  }

  // 重新渲染时清空选择
  selectedRows.clear();

  var cols = getDisplayColumns();
  var metaSuffix = lastFieldSchema
    ? "（含空字段，按表格列顺序）"
    : "（schema 加载中…）";

  var html = '<div id="selectionToolbar" class="selection-toolbar">'
           +   '<span><span id="selectionCount">0</span> 行已选中</span>'
           +   '<div class="btn-group">'
           +     '<button onclick="copySelectedIds()">📎 复制选中的 recordId</button>'
           +     '<button class="secondary" onclick="useSelectedIds()">↩ 用选中 ID 填到 update/delete</button>'
           +     '<button class="secondary" onclick="clearSelection()">取消选择</button>'
           +   '</div>'
           + '</div>';

  html += '<div class="table-meta">共 ' + lastRecords.length + ' 条记录 · '
        + cols.length + ' 个字段 ' + metaSuffix
        + ' · <span style="color:var(--text-3);">点击=单选 · Cmd/Ctrl+点击=多选 · Shift+点击=区间 · 上下拖=多行 · 左右拖=横滚</span>'
        + '</div>';
  html += '<table><thead><tr>';
  html += '<th>recordId</th><th>createdAt<div class="th-type">日期</div></th><th>updatedAt<div class="th-type">日期</div></th>';
  cols.forEach(function(c) {
    var typeLabel = fieldTypeLabel(c.type);
    html += '<th>' + escapeHtml(c.header)
         + (typeLabel ? '<div class="th-type">' + escapeHtml(typeLabel) + '</div>' : '')
         + '</th>';
  });
  html += '</tr></thead><tbody>';

  lastRecords.forEach(function(r, idx) {
    html += '<tr data-row-idx="' + idx + '">';
    html += '<td class="id-col">' + escapeHtml(r.recordId || "") + '</td>';
    html += '<td class="id-col">' + escapeHtml(formatTimestamp(r.createdAt)) + '</td>';
    html += '<td class="id-col">' + escapeHtml(formatTimestamp(r.updatedAt)) + '</td>';
    cols.forEach(function(c) {
      var v = r.fields ? r.fields[c.key] : "";
      var text = formatCellValue(v, c.type);
      html += '<td title="' + escapeHtml(text) + '">' + escapeHtml(text) + '</td>';
    });
    html += '</tr>';
  });
  html += '</tbody></table>';
  container.innerHTML = html;
  setupTableSelection();
}

function setupTableSelection() {
  var tableViewEl = document.getElementById("tableView");
  var tbody = document.querySelector("#tableView table tbody");
  if (!tbody || !tableViewEl) return;

  tbody.addEventListener("mousedown", function(e) {
    var tr = e.target.closest("tr[data-row-idx]");
    if (!tr) return;
    e.preventDefault();
    var idx = parseInt(tr.dataset.rowIdx, 10);

    // 记录起点 + 当前滚动位置（pan 时用）
    _dragStartX = e.clientX;
    _dragStartY = e.clientY;
    _initialScrollLeft = tableViewEl.scrollLeft;
    _dragMode = null;

    // 立即做单击行为（之后如果决定是 pan 会撤销）
    if (e.shiftKey && lastClickedRowIdx !== -1) {
      var start = Math.min(lastClickedRowIdx, idx);
      var end = Math.max(lastClickedRowIdx, idx);
      for (var i = start; i <= end; i++) selectedRows.add(i);
    } else if (e.metaKey || e.ctrlKey) {
      if (selectedRows.has(idx)) selectedRows.delete(idx);
      else selectedRows.add(idx);
    } else {
      selectedRows.clear();
      selectedRows.add(idx);
    }

    lastClickedRowIdx = idx;
    dragStartRowIdx = idx;
    isDraggingSelection = true;
    updateSelectionUI();
  });
}

// 全局 mousemove：在拖动中实时判断方向并执行 pan / select
document.addEventListener("mousemove", function(e) {
  if (!isDraggingSelection) return;
  var tableViewEl = document.getElementById("tableView");
  if (!tableViewEl) return;

  var dx = e.clientX - _dragStartX;
  var dy = e.clientY - _dragStartY;

  // 决定模式：第一次明显移动决定方向
  if (_dragMode === null) {
    if (Math.abs(dx) > _DRAG_THRESHOLD || Math.abs(dy) > _DRAG_THRESHOLD) {
      if (Math.abs(dx) > Math.abs(dy)) {
        _dragMode = "pan";
        // 撤销初始单击造成的多选（保留刚点的那一行）
        selectedRows.clear();
        selectedRows.add(dragStartRowIdx);
        updateSelectionUI();
        document.body.style.cursor = "grabbing";
        // 防止文字选中
        if (window.getSelection) window.getSelection().removeAllRanges();
      } else {
        _dragMode = "select";
      }
    }
  }

  if (_dragMode === "pan") {
    tableViewEl.scrollLeft = _initialScrollLeft - dx;
  } else if (_dragMode === "select") {
    // 找到当前鼠标下的行
    var el = document.elementFromPoint(e.clientX, e.clientY);
    var tr = el && el.closest && el.closest("#tableView tr[data-row-idx]");
    if (tr) {
      var idx = parseInt(tr.dataset.rowIdx, 10);
      var s = Math.min(dragStartRowIdx, idx);
      var en = Math.max(dragStartRowIdx, idx);
      selectedRows.clear();
      for (var i = s; i <= en; i++) selectedRows.add(i);
      updateSelectionUI();
    }
  }
});

document.addEventListener("mouseup", function() {
  if (isDraggingSelection) {
    isDraggingSelection = false;
    _dragMode = null;
    document.body.style.cursor = "";
  }
});

function updateSelectionUI() {
  document.querySelectorAll("#tableView tbody tr[data-row-idx]").forEach(function(tr) {
    var idx = parseInt(tr.dataset.rowIdx, 10);
    tr.classList.toggle("row-selected", selectedRows.has(idx));
  });
  var toolbar = document.getElementById("selectionToolbar");
  var countEl = document.getElementById("selectionCount");
  if (toolbar) toolbar.classList.toggle("active", selectedRows.size > 0);
  if (countEl) countEl.textContent = selectedRows.size;
}

function getSelectedRecordIds() {
  if (!lastRecords) return [];
  var ids = [];
  // 按行索引顺序排
  Array.from(selectedRows).sort(function(a, b) { return a - b; }).forEach(function(idx) {
    var r = lastRecords[idx];
    if (r && r.recordId) ids.push(r.recordId);
  });
  return ids;
}

function copySelectedIds() {
  var ids = getSelectedRecordIds();
  if (ids.length === 0) { flash("没有选中行"); return; }
  var text = ids.join(",");
  if (navigator.clipboard && navigator.clipboard.writeText) {
    navigator.clipboard.writeText(text).then(function() {
      flash("✓ 已复制 " + ids.length + " 个 recordId 到剪贴板");
    }).catch(function() { _fallbackCopy(text, ids.length); });
  } else {
    _fallbackCopy(text, ids.length);
  }
}

function _fallbackCopy(text, count) {
  var ta = document.createElement("textarea");
  ta.value = text; ta.style.position = "fixed"; ta.style.opacity = "0";
  document.body.appendChild(ta); ta.select();
  try { document.execCommand("copy"); flash("✓ 已复制 " + count + " 个 recordId"); }
  catch (e) { flash("复制失败"); }
  document.body.removeChild(ta);
}

// 把选中的 recordId 自动填到 update/delete 表单的 recordIds 字段
// 如果当前不是 update/delete，先切到 delete（最常用）
function useSelectedIds() {
  var ids = getSelectedRecordIds();
  if (ids.length === 0) { flash("没有选中行"); return; }
  var op = document.getElementById("operation").value;
  if (op !== "delete" && op !== "update") {
    if (!confirm("当前操作是「" + opLabel(op) + "」，没有 recordIds 字段。\n是否切换到「删除记录」并自动填入 " + ids.length + " 个 recordId？")) {
      return;
    }
    document.getElementById("operation").value = "delete";
    onOperationChange();
    op = "delete";
  }
  // 等下一帧 DOM 更新完
  setTimeout(function() {
    var el = document.getElementById("param_recordIds");
    if (el) {
      el.value = ids.join(",");
      el.scrollIntoView({behavior: "smooth", block: "center"});
      el.style.transition = "background 0.4s";
      el.style.background = "var(--primary-soft)";
      setTimeout(function() { el.style.background = ""; }, 800);
      flash("✓ 已填入 " + ids.length + " 个 recordId 到 " + opLabel(op) + " 的 recordIds 字段");
    } else {
      flash("找不到 recordIds 输入框");
    }
  }, 0);
}

function clearSelection() {
  selectedRows.clear();
  updateSelectionUI();
}

// ---------- 按内容搜索 recordId 小部件 ----------
var _lastFilterSearchIds = [];

// 可视化构建器状态
var filterConditions = [{ field: "", operator: "=", value: "" }];
var filterConnector = "AND";  // AND / OR
var filterMode = "visual";    // visual / formula

// 操作符定义：每个有 label（中文）和 template（生成 AITable 公式片段）
// {F} 替换字段名（已带大括号），{V} 替换值
var FILTER_OPS = [
  { value: "=",        label: "等于",       template: '{F}="{V}"',         needsValue: true,  numeric: false },
  { value: "!=",       label: "不等于",     template: '{F}!="{V}"',        needsValue: true,  numeric: false },
  { value: "contains", label: "包含",       template: 'FIND("{V}",{F})',   needsValue: true,  numeric: false },
  { value: ">=",       label: "大于等于",   template: '{F}>={V}',          needsValue: true,  numeric: true  },
  { value: "<=",       label: "小于等于",   template: '{F}<={V}',          needsValue: true,  numeric: true  },
  { value: ">",        label: "大于",       template: '{F}>{V}',           needsValue: true,  numeric: true  },
  { value: "<",        label: "小于",       template: '{F}<{V}',           needsValue: true,  numeric: true  },
  { value: "empty",    label: "为空",       template: '{F}=""',            needsValue: false, numeric: false },
  { value: "notempty", label: "不为空",     template: '{F}!=""',           needsValue: false, numeric: false }
];

function buildFilterSearchWidget() {
  var wrap = document.createElement("div");
  wrap.className = "filter-search";
  wrap.innerHTML =
    '<button type="button" class="filter-search-toggle" onclick="toggleFilterSearch(this)">' +
      '🔍 不知道 recordId？点这里按内容搜索' +
    '</button>' +
    '<div class="filter-search-panel" id="filterSearchPanel" style="display:none;">' +
      '<div class="filter-mode-tabs">' +
        '<button id="filterModeVisual" class="active" onclick="switchFilterMode(\'visual\')">🧱 可视化构建</button>' +
        '<button id="filterModeFormula" onclick="switchFilterMode(\'formula\')">📝 公式</button>' +
      '</div>' +
      '<div id="filterVisualBuilder"></div>' +
      '<div id="filterFormulaBuilder" style="display:none;">' +
        '<div class="row">' +
          '<input type="text" id="filterFormulaInput" ' +
                 'placeholder=\'例：{标题}="赛季问卷"\' ' +
                 'onkeydown="if(event.key===\'Enter\'){event.preventDefault();executeFilterSearch();}">' +
        '</div>' +
        '<div class="examples">' +
          '常用例子：' +
          '<code>{标题}="赛季问卷"</code> ' +
          '<code>{重要度}&gt;=4</code> ' +
          '<code>AND({状态}="进行中",{负责人}="张三")</code> ' +
          '<code>{标题}!=""</code>' +
        '</div>' +
      '</div>' +
      '<div class="actions" style="margin-top:10px;">' +
        '<button class="small" onclick="executeFilterSearch()">🔍 查找匹配的记录</button>' +
      '</div>' +
      '<div id="filterFormulaPreview" class="filter-formula-preview" style="display:none;"></div>' +
      '<div id="filterSearchResults" class="filter-search-results"></div>' +
    '</div>';
  return wrap;
}

function toggleFilterSearch(btn) {
  var panel = document.getElementById("filterSearchPanel");
  var open = panel.style.display === "none";
  panel.style.display = open ? "" : "none";
  btn.textContent = open ? "🔍 收起搜索面板" : "🔍 不知道 recordId？点这里按内容搜索";
  if (open) {
    // 第一次打开：重置条件 + 拉 schema
    if (filterConditions.length === 1 && !filterConditions[0].field) {
      // 保持空白条件
    }
    renderFilterVisualBuilder();
  }
}

function switchFilterMode(mode) {
  filterMode = mode;
  document.getElementById("filterModeVisual").classList.toggle("active", mode === "visual");
  document.getElementById("filterModeFormula").classList.toggle("active", mode === "formula");
  document.getElementById("filterVisualBuilder").style.display = mode === "visual" ? "" : "none";
  document.getElementById("filterFormulaBuilder").style.display = mode === "formula" ? "" : "none";
  if (mode === "visual") {
    renderFilterVisualBuilder();
  } else {
    // 切到公式模式：把可视化生成的公式放进输入框
    var formula = buildFormulaFromConditions();
    var inp = document.getElementById("filterFormulaInput");
    if (inp && formula) inp.value = formula;
  }
  updateFormulaPreview();
}

function renderFilterVisualBuilder() {
  var container = document.getElementById("filterVisualBuilder");
  if (!container) return;

  // 没 schema 就先拉一下
  if (!lastFieldSchema) {
    container.innerHTML = '<div style="color:var(--text-3);font-size:12px;padding:8px;">⏳ 加载字段列表...</div>';
    fetchSchemaForFilter().then(function(ok) {
      if (ok) renderFilterVisualBuilder();
      else container.innerHTML = '<div style="color:var(--danger);font-size:12px;padding:8px;">无法加载字段列表，请检查 Token / 表格 ID 是否正确</div>';
    });
    return;
  }

  container.innerHTML = "";

  // 多条件时显示连接符选择
  if (filterConditions.length > 1) {
    var conn = document.createElement("div");
    conn.className = "filter-connector";
    conn.innerHTML = '所有条件用 <select onchange="filterConnector=this.value;updateFormulaPreview();">' +
      '<option value="AND"' + (filterConnector === "AND" ? " selected" : "") + '>且 (AND，全部满足)</option>' +
      '<option value="OR"' + (filterConnector === "OR" ? " selected" : "") + '>或 (OR，任一满足)</option>' +
      '</select> 连接';
    container.appendChild(conn);
  }

  // 条件行
  filterConditions.forEach(function(cond, idx) {
    var row = document.createElement("div");
    row.className = "filter-cond-row";

    // 字段下拉
    var fieldSelect = document.createElement("select");
    fieldSelect.innerHTML = '<option value="">— 选择字段 —</option>';
    lastFieldSchema.forEach(function(f) {
      var opt = document.createElement("option");
      opt.value = f.name;
      opt.textContent = f.name + " (" + (fieldTypeLabel(f.type) || "?") + ")";
      if (cond.field === f.name) opt.selected = true;
      fieldSelect.appendChild(opt);
    });
    fieldSelect.onchange = function() {
      filterConditions[idx].field = fieldSelect.value;
      updateFormulaPreview();
    };
    row.appendChild(fieldSelect);

    // 操作符下拉
    var opSelect = document.createElement("select");
    FILTER_OPS.forEach(function(op) {
      var opt = document.createElement("option");
      opt.value = op.value;
      opt.textContent = op.label;
      if (cond.operator === op.value) opt.selected = true;
      opSelect.appendChild(opt);
    });
    opSelect.onchange = function() {
      filterConditions[idx].operator = opSelect.value;
      renderFilterVisualBuilder();  // 重渲染：可能从有值变成无值
    };
    row.appendChild(opSelect);

    // 值输入（如果操作符需要值）
    var opDef = FILTER_OPS.find(function(o) { return o.value === cond.operator; });
    if (opDef && opDef.needsValue) {
      var valInput = document.createElement("input");
      valInput.type = "text";
      valInput.placeholder = opDef.numeric ? "数字" : "值";
      valInput.value = cond.value || "";
      valInput.oninput = function() {
        filterConditions[idx].value = valInput.value;
        updateFormulaPreview();
      };
      valInput.onkeydown = function(e) {
        if (e.key === "Enter") { e.preventDefault(); executeFilterSearch(); }
      };
      row.appendChild(valInput);
    } else {
      var noVal = document.createElement("div");
      noVal.className = "empty-value";
      noVal.textContent = "—";
      row.appendChild(noVal);
    }

    // 删除按钮
    var del = document.createElement("button");
    del.type = "button";
    del.className = "del-btn";
    del.textContent = "✕";
    del.disabled = filterConditions.length === 1;
    del.style.opacity = filterConditions.length === 1 ? "0.3" : "1";
    del.onclick = function() {
      if (filterConditions.length > 1) {
        filterConditions.splice(idx, 1);
        renderFilterVisualBuilder();
      }
    };
    row.appendChild(del);

    container.appendChild(row);
  });

  // 添加条件按钮
  var addBtn = document.createElement("button");
  addBtn.type = "button";
  addBtn.className = "ghost small";
  addBtn.textContent = "+ 添加条件";
  addBtn.style.marginTop = "4px";
  addBtn.onclick = function() {
    filterConditions.push({ field: "", operator: "=", value: "" });
    renderFilterVisualBuilder();
  };
  container.appendChild(addBtn);

  updateFormulaPreview();
}

function fetchSchemaForFilter() {
  return new Promise(function(resolve) {
    var token = document.getElementById("token").value.trim();
    var baseUrl = document.getElementById("baseUrl").value.trim();
    var dstInput = document.getElementById("param_datasheetId");
    var datasheetId = dstInput ? dstInput.value.trim() : "";
    if (!token || !baseUrl || !datasheetId) {
      resolve(false); return;
    }
    fetch("/api/call", {
      method: "POST", headers: {"Content-Type": "application/json"},
      body: JSON.stringify({
        op: "get_fields", token: token, baseUrl: baseUrl, datasheetId: datasheetId,
        params: { datasheetId: datasheetId }
      })
    }).then(function(r) { return r.json(); })
      .then(function(data) {
        if (data && data.response && data.response.data
            && Array.isArray(data.response.data.fields)) {
          lastFieldSchema = data.response.data.fields;
          resolve(true);
        } else { resolve(false); }
      })
      .catch(function() { resolve(false); });
  });
}

// 把可视化条件转成 AITable 筛选公式
function buildFormulaFromConditions() {
  var parts = filterConditions
    .filter(function(c) { return c.field; })
    .map(function(c) {
      var op = null;
      for (var i = 0; i < FILTER_OPS.length; i++) {
        if (FILTER_OPS[i].value === c.operator) { op = FILTER_OPS[i]; break; }
      }
      if (!op) return null;
      var template = op.template;
      var value = String(c.value || "").replace(/"/g, '\\"');
      return template.replace(/\{F\}/g, "{" + c.field + "}").replace(/\{V\}/g, value);
    })
    .filter(Boolean);
  if (parts.length === 0) return "";
  if (parts.length === 1) return parts[0];
  return filterConnector + "(" + parts.join(",") + ")";
}

function updateFormulaPreview() {
  var preview = document.getElementById("filterFormulaPreview");
  if (!preview) return;
  if (filterMode !== "visual") {
    preview.style.display = "none";
    return;
  }
  var formula = buildFormulaFromConditions();
  if (!formula) {
    preview.style.display = "none";
    return;
  }
  preview.style.display = "block";
  preview.innerHTML = '<span class="label">→ 生成的公式</span>' + escapeHtml(formula);
}

function executeFilterSearch() {
  var formula;
  if (filterMode === "visual") {
    formula = buildFormulaFromConditions();
  } else {
    var inp = document.getElementById("filterFormulaInput");
    formula = inp ? inp.value.trim() : "";
  }

  var resultsEl = document.getElementById("filterSearchResults");
  if (!formula) {
    resultsEl.classList.add("active", "error");
    resultsEl.textContent = "⚠ 请先填好筛选条件" + (filterMode === "visual" ? "（至少选一个字段）" : "");
    return;
  }

  var token = document.getElementById("token").value.trim();
  var baseUrl = document.getElementById("baseUrl").value.trim();
  var dstInput = document.getElementById("param_datasheetId");
  var datasheetId = dstInput ? dstInput.value.trim() : "";

  if (!token || !baseUrl || !datasheetId) {
    resultsEl.classList.add("active", "error");
    resultsEl.textContent = "⚠ 请先填好 顶部 Token / 接口地址 / 表格 ID";
    return;
  }

  resultsEl.classList.add("active");
  resultsEl.classList.remove("error");
  resultsEl.innerHTML = "⏳ 查询中...";

  fetch("/api/call", {
    method: "POST", headers: {"Content-Type": "application/json"},
    body: JSON.stringify({
      op: "get_records", token: token, baseUrl: baseUrl, datasheetId: datasheetId,
      params: {
        datasheetId: datasheetId,
        filterByFormula: formula,
        fieldKey: "name",
        pageSize: "100"
      }
    })
  }).then(function(r) { return r.json(); })
    .then(function(data) {
      if (data.status >= 400) {
        resultsEl.classList.add("error");
        var errText = "查询失败 HTTP " + data.status;
        try { errText += "\n" + JSON.stringify(data.response); }
        catch (e) {}
        resultsEl.textContent = errText;
        return;
      }
      var records = extractRecords(data.response) || [];
      var ids = records.map(function(r) { return r.recordId; }).filter(Boolean);
      _lastFilterSearchIds = ids;

      if (ids.length === 0) {
        resultsEl.innerHTML = '<div style="color:var(--text-3);">没有匹配的记录</div>';
        return;
      }

      var op = document.getElementById("operation").value;
      var apiLimit = 10;
      var html = '<div>找到 <b>' + ids.length + '</b> 条匹配记录' +
                 (ids.length > apiLimit ? '（' + opLabel(op).toUpperCase() + ' 一次最多 ' + apiLimit + ' 条，超出会被截断）' : '') + '</div>';
      html += '<ul>';
      records.slice(0, 10).forEach(function(r) {
        var preview = "";
        if (r.fields) {
          var firstKey = Object.keys(r.fields)[0];
          if (firstKey) preview = " · " + firstKey + ": " + formatCellValue(r.fields[firstKey]);
        }
        html += '<li>' + escapeHtml(r.recordId) + escapeHtml(preview) + '</li>';
      });
      if (ids.length > 10) html += '<li>... +' + (ids.length - 10) + ' 条</li>';
      html += '</ul>';
      html += '<div class="actions">';
      var n = Math.min(apiLimit, ids.length);
      var btnLabel;
      if (op === "update") {
        btnLabel = '✓ 用这 ' + n + ' 个 recordId 批量生成更新记录（沿用第 1 条的字段为模板）';
      } else {
        btnLabel = '✓ 把这 ' + n + ' 个 ID 填入 recordIds';
      }
      html += '<button class="small" onclick="useFilterResults()">' + btnLabel + '</button>';
      html += '</div>';
      resultsEl.innerHTML = html;
    })
    .catch(function(e) {
      resultsEl.classList.add("error");
      resultsEl.textContent = "✗ 失败: " + e.message;
    });
}

function useFilterResults() {
  var op = document.getElementById("operation").value;
  if (op === "update") {
    applyFilterResultsToUpdate();
  } else {
    applyFilterResultsToDelete();
  }
}

function applyFilterResultsToDelete() {
  var ids = _lastFilterSearchIds.slice(0, 10);
  if (ids.length === 0) return;
  var el = document.getElementById("param_recordIds");
  if (!el) { flash("找不到 recordIds 字段"); return; }
  el.value = ids.join(",");
  el.style.transition = "background 0.4s";
  el.style.background = "var(--primary-soft)";
  setTimeout(function() { el.style.background = ""; }, 800);
  flash("✓ 已填入 " + ids.length + " 个 recordId" +
        (_lastFilterSearchIds.length > 10 ? "（共 " + _lastFilterSearchIds.length + " 条匹配，已截到前 10 条）" : ""));
}

function applyFilterResultsToUpdate() {
  var ids = _lastFilterSearchIds.slice(0, 10);
  if (ids.length === 0) return;

  // 取第 1 条记录的"已填了字段名"的字段作为模板
  var template = [];
  if (recordsState.update && recordsState.update.length > 0) {
    template = (recordsState.update[0].fields || []).filter(function(f) { return f.name; });
  }

  if (template.length === 0) {
    if (!confirm(
      "⚠ 当前可视化编辑器里第 1 条记录还没填任何字段（作为模板）。\n\n" +
      "如果继续，会生成 " + ids.length + " 条「只有 recordId、没有字段」的空记录，发请求会失败。\n\n" +
      "建议先取消，回去在「第 1 条」里填好要修改的字段（比如 标题=新标题），\n" +
      "再点「批量生成」，工具会把这些字段复制到所有匹配的 recordId 上。\n\n" +
      "确定要现在继续生成空记录吗？"
    )) {
      return;
    }
  }

  // 用模板字段克隆 N 条
  recordsState.update = ids.map(function(id) {
    return {
      recordId: id,
      fields: template.map(function(f) {
        return { name: f.name, value: f.value, type: f.type };
      })
    };
  });

  // 重新渲染编辑器
  var ed = document.getElementById("records_editor_update");
  if (ed) renderRecordsEditor(ed, "update");

  var msg = "✓ 已批量生成 " + ids.length + " 条更新记录";
  if (template.length > 0) {
    msg += "，每条沿用了 " + template.length + " 个模板字段：" +
           template.map(function(f) { return f.name; }).join("、");
  }
  if (_lastFilterSearchIds.length > 10) {
    msg += "（共 " + _lastFilterSearchIds.length + " 条匹配，已截到前 10 条）";
  }
  flash(msg);

  // 滚动到记录编辑器
  if (ed) ed.scrollIntoView({behavior: "smooth", block: "start"});
}

function switchView(view) {
  var btns = document.querySelectorAll("#responseToolbar .view-toggle button");
  btns.forEach(function(b) {
    b.classList.toggle("active", b.dataset.view === view);
  });
  document.getElementById("result").style.display = view === "json" ? "" : "none";
  document.getElementById("tableView").style.display = view === "table" ? "" : "none";
  if (view === "table") renderTableView();
}

function csvEscape(s) {
  s = String(s == null ? "" : s);
  if (/[",\r\n]/.test(s)) return '"' + s.replace(/"/g, '""') + '"';
  return s;
}

function downloadBlob(content, filename, mimeType) {
  var blob = new Blob([content], { type: mimeType });
  var url = URL.createObjectURL(blob);
  var a = document.createElement("a");
  a.href = url; a.download = filename;
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
  setTimeout(function() { URL.revokeObjectURL(url); }, 1000);
}

function nowStamp() {
  var d = new Date();
  function pad(n) { return n < 10 ? "0" + n : n; }
  return d.getFullYear() + pad(d.getMonth()+1) + pad(d.getDate()) + "_"
       + pad(d.getHours()) + pad(d.getMinutes()) + pad(d.getSeconds());
}

function downloadCSV() {
  if (!lastRecords) {
    alert("当前响应不是查询记录结果，无法导出表格。\n\n请先用「查询记录」操作得到带 records 的响应。");
    return;
  }
  var cols = getDisplayColumns();
  var headers = ["recordId", "createdAt", "updatedAt"].concat(cols.map(function(c) { return c.header; }));
  var lines = [headers.map(csvEscape).join(",")];
  lastRecords.forEach(function(r) {
    var row = [
      r.recordId || "",
      formatTimestamp(r.createdAt),
      formatTimestamp(r.updatedAt)
    ];
    cols.forEach(function(c) {
      row.push(formatCellValue(r.fields ? r.fields[c.key] : "", c.type));
    });
    lines.push(row.map(csvEscape).join(","));
  });
  // BOM 让 Excel 正确识别 UTF-8 中文
  var csv = "﻿" + lines.join("\r\n");
  downloadBlob(csv, "aitable_records_" + nowStamp() + ".csv", "text/csv;charset=utf-8");
}

function downloadJSON() {
  if (!lastResponse) { alert("暂无响应数据"); return; }
  var content = JSON.stringify(lastResponse.response, null, 2);
  downloadBlob(content, "aitable_response_" + nowStamp() + ".json", "application/json;charset=utf-8");
}

function sendRequest() {
  var op = document.getElementById("operation").value;
  var token = document.getElementById("token").value.trim();
  var baseUrl = document.getElementById("baseUrl").value;

  if (!token) { showResult("⚠ 请填写 API 密钥", true); return; }

  var params = {};
  var formFields = FORMS[op] || [];
  for (var i = 0; i < formFields.length; i++) {
    var f = formFields[i];
    params[f.key] = getParamValue(f.key, op);
  }

  if (op !== "get_spaces" && !params.datasheetId) {
    showResult("⚠ 请填写表格 ID", true); return;
  }
  var datasheetId = params.datasheetId || "";

  // CREATE/UPDATE 在可视化模式下：检查重复字段名
  if ((op === "create" || op === "update") && !jsonModeState[op]) {
    var dupSummary = [];
    (recordsState[op] || []).forEach(function(rec, idx) {
      var d = detectDuplicateFields(rec);
      if (d.names.length > 0) {
        dupSummary.push("第 " + (idx + 1) + " 条：" + d.names.join("、"));
      }
    });
    if (dupSummary.length > 0) {
      if (!confirm(
        "⚠ 检测到字段名重复：\n\n" + dupSummary.join("\n") + "\n\n" +
        "JSON 对象的 key 必须唯一，所以同名字段会互相覆盖（只保留最后一个值）。\n\n" +
        "如果你想创建多条记录，请先取消，把重复的字段拆到不同记录里（用「+ 添加一条记录」）。\n\n" +
        "确定要继续按现状发送吗？"
      )) {
        return;
      }
    }
  }

  showLoading("发送 " + opLabel(op) + " 请求");
  document.getElementById("responseToolbar").style.display = "none";
  document.getElementById("tableView").style.display = "none";
  document.getElementById("result").style.display = "";

  // 记下这次的环境信息，refetch / 打开外链时用
  lastDatasheetId = datasheetId;
  lastBaseUrl = baseUrl;
  lastToken = token;
  lastFieldSchema = null;
  // 记下这次用的 fieldKey（默认 name）
  lastFieldKey = (params.fieldKey === "id") ? "id" : "name";

  fetch("/api/call", {
    method: "POST",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify({op: op, token: token, datasheetId: datasheetId, baseUrl: baseUrl, params: params})
  }).then(function(resp) { return resp.json(); })
    .then(function(data) {
      hideLoading();
      lastResponse = data;
      lastRecords = extractRecords(data.response);
      renderResponse();
      // 后台静默拉 schema（不阻塞渲染），拉到了再刷新表格视图
      maybeFetchSchema();
    }).catch(function(e) {
      hideLoading();
      lastResponse = null; lastRecords = null;
      document.getElementById("responseToolbar").style.display = "none";
      showResult("✗ 请求失败: " + e.message, true);
    });
}

// 把 lastResponse 渲染到界面（独立出来便于 refetch 复用）
function renderResponse() {
  var data = lastResponse;
  if (!data) return;
  var text = "➡  " + data.request.method + " " + data.request.url + "\n";
  if (data.request.body) {
    text += "Body:\n" + JSON.stringify(data.request.body, null, 2) + "\n";
  }
  text += "──────────────────────────────────────────────\n";
  text += "HTTP " + data.status + "\n";
  text += typeof data.response === "string"
    ? data.response : JSON.stringify(data.response, null, 2);
  showResult(text, data.status >= 400);

  document.getElementById("responseToolbar").style.display = "flex";
  var tableBtn = document.querySelector('#responseToolbar .view-toggle button[data-view="table"]');
  if (tableBtn) {
    tableBtn.disabled = !lastRecords;
    tableBtn.style.opacity = lastRecords ? "" : "0.4";
    tableBtn.style.cursor = lastRecords ? "" : "not-allowed";
    tableBtn.title = lastRecords
      ? ("可切换到表格视图（" + lastRecords.length + " 条记录）")
      : "当前响应不是查询记录结果";
  }

  // 「重新查询本次记录」按钮：当响应里包含可用 recordId 时显示
  var btnRefetch = document.getElementById("btnRefetch");
  var hasIds = lastRecords && lastRecords.length > 0
            && lastRecords.some(function(r) { return r.recordId; });
  btnRefetch.style.display = (hasIds && lastDatasheetId) ? "" : "none";

  // 「在 APITable 中打开」按钮：当有 datasheetId 时显示
  var btnOpen = document.getElementById("btnOpenAitable");
  btnOpen.style.display = lastDatasheetId ? "" : "none";

  if (lastRecords && lastRecords.length > 0) {
    switchView("table");
  } else {
    switchView("json");
  }
}

// 用上次响应里的 recordId 列表，重新查一次记录 + 表格字段结构
// 这样表格视图能显示所有字段（包括空字段），按表格的列顺序展示
function refetchFullRecords() {
  if (!lastRecords || lastRecords.length === 0) {
    flash("当前没有记录数据，无法重新查询"); return;
  }
  var ids = lastRecords.map(function(r) { return r.recordId; }).filter(Boolean);
  if (ids.length === 0) { flash("没有可用的 recordId"); return; }
  if (!lastDatasheetId) { flash("没有表格 ID，无法重新查询"); return; }
  if (!lastToken) { flash("没有 Token"); return; }

  showLoading("重新查询 " + ids.length + " 条记录的完整字段");

  var recordsReq = fetch("/api/call", {
    method: "POST", headers: {"Content-Type": "application/json"},
    body: JSON.stringify({
      op: "get_records", token: lastToken, baseUrl: lastBaseUrl, datasheetId: lastDatasheetId,
      params: {
        datasheetId: lastDatasheetId, recordIds: ids.join(","), fieldKey: "name"
      }
    })
  }).then(function(r) { return r.json(); });

  var fieldsReq = fetch("/api/call", {
    method: "POST", headers: {"Content-Type": "application/json"},
    body: JSON.stringify({
      op: "get_fields", token: lastToken, baseUrl: lastBaseUrl, datasheetId: lastDatasheetId,
      params: { datasheetId: lastDatasheetId }
    })
  }).then(function(r) { return r.json(); });

  Promise.all([recordsReq, fieldsReq]).then(function(results) {
    hideLoading();
    var recordsResp = results[0];
    var fieldsResp = results[1];

    if (fieldsResp && fieldsResp.response && fieldsResp.response.data
        && Array.isArray(fieldsResp.response.data.fields)) {
      lastFieldSchema = fieldsResp.response.data.fields;
    } else {
      lastFieldSchema = null;
    }

    lastResponse = recordsResp;
    lastRecords = extractRecords(recordsResp.response);
    lastFieldKey = "name";
    renderResponse();
    var msg = "✓ 已重新查询 " + ids.length + " 条记录";
    if (lastFieldSchema) msg += "，含全部 " + lastFieldSchema.length + " 个字段（按表格列顺序）";
    flash(msg);
  }).catch(function(e) {
    hideLoading();
    showResult("✗ 重新查询失败: " + e.message, true);
  });
}

// 不管刚才做了什么操作，都可以一键拉整张表（最多 100 条）
// 这是「create/update/delete 后想看全表」的真正解法 —— 因为这些 API 永远只返回受影响的那几条
function queryAllRecords() {
  // 「查询整张表」是个独立操作 —— 不依赖之前的请求，从当前表单/通用配置读取
  var token = document.getElementById("token").value.trim();
  var baseUrl = document.getElementById("baseUrl").value.trim();
  var dstInput = document.getElementById("param_datasheetId");
  var datasheetId = dstInput ? dstInput.value.trim() : "";

  if (!token) { alert("请先填写顶部的「API 密钥」"); return; }
  if (!baseUrl) { alert("请先填写顶部的「接口地址」"); return; }
  if (!datasheetId) { alert("请先在参数区填写「表格 ID」"); return; }

  // 检查 viewId（不同操作字段都叫 viewId）
  var viewIdInput = document.getElementById("param_viewId");
  var viewId = viewIdInput ? viewIdInput.value.trim() : "";

  if (!viewId) {
    var go = confirm(
      "提示：还没填写「视图 ID（viewId）」。\n\n" +
      "不填的话，会按表格的默认视图查询整张表（通常包含所有字段、所有记录）。\n" +
      "如果只想看某个特定视图的数据，请取消后回到参数区填上 viewId 再试。\n\n" +
      "点「确定」继续按默认视图查询，点「取消」回去填 viewId。"
    );
    if (!go) {
      if (viewIdInput) viewIdInput.focus();
      return;
    }
  }

  // 把这次的环境信息记下来，让 refetch / openInAitable 等仍能工作
  lastDatasheetId = datasheetId;
  lastBaseUrl = baseUrl;
  lastToken = token;
  lastFieldSchema = null;

  showLoading("查询整张表（最多 100 条）" + (viewId ? "·视图 " + viewId : "·默认视图") + " + 字段结构");

  var recordsParams = {
    datasheetId: datasheetId,
    fieldKey: "name",
    pageSize: "100"
  };
  if (viewId) recordsParams.viewId = viewId;

  var recordsReq = fetch("/api/call", {
    method: "POST", headers: {"Content-Type": "application/json"},
    body: JSON.stringify({
      op: "get_records", token: token, baseUrl: baseUrl, datasheetId: datasheetId,
      params: recordsParams
    })
  }).then(function(r) { return r.json(); });

  // 顺便拉 schema，让所有列都能展示
  var fieldsReq = fetch("/api/call", {
    method: "POST", headers: {"Content-Type": "application/json"},
    body: JSON.stringify({
      op: "get_fields", token: token, baseUrl: baseUrl, datasheetId: datasheetId,
      params: { datasheetId: datasheetId }
    })
  }).then(function(r) { return r.json(); });

  Promise.all([recordsReq, fieldsReq]).then(function(results) {
    hideLoading();
    var recordsResp = results[0], fieldsResp = results[1];
    if (fieldsResp && fieldsResp.response && fieldsResp.response.data
        && Array.isArray(fieldsResp.response.data.fields)) {
      lastFieldSchema = fieldsResp.response.data.fields;
    }
    lastResponse = recordsResp;
    lastRecords = extractRecords(recordsResp.response);
    lastFieldKey = "name";
    renderResponse();
    var n = lastRecords ? lastRecords.length : 0;
    flash("✓ 已查询整张表，共 " + n + " 条记录"
          + (n === 100 ? "（达到本次单页上限，可在「查询记录」操作中翻页）" : ""));
  }).catch(function(e) {
    hideLoading();
    showResult("✗ 查询失败: " + e.message, true);
  });
}

// ==========================================================
//   Excel / CSV 批量导入
// ==========================================================
// 解析的文件数据
var importFile = null;       // { name, headers: [], rows: [] }
var importMapping = {};      // { fileColIdx: aitableFieldName  ('' = 忽略) }
var importSource = "browser"; // "browser" 或 "path"

function switchImportSource(src) {
  importSource = src;
  var bBtn = document.getElementById("importSrcBrowserBtn");
  var pBtn = document.getElementById("importSrcPathBtn");
  bBtn.style.background = src === "browser" ? "white" : "transparent";
  bBtn.style.color = src === "browser" ? "var(--primary)" : "var(--text-2)";
  pBtn.style.background = src === "path" ? "white" : "transparent";
  pBtn.style.color = src === "path" ? "var(--primary)" : "var(--text-2)";
  document.getElementById("importSrcBrowser").style.display = src === "browser" ? "" : "none";
  document.getElementById("importSrcPath").style.display = src === "path" ? "" : "none";
}

function testImportPath() {
  var path = document.getElementById("importFilePath").value.trim();
  if (!path) { alert("请填写文件路径"); return; }
  var info = document.getElementById("importPathInfo");
  info.innerHTML = "⏳ 读取中...";
  fetch("/api/read-file", {
    method: "POST", headers: {"Content-Type": "application/json"},
    body: JSON.stringify({ path: path })
  }).then(function(r) { return r.json(); })
    .then(function(data) {
      if (data.error) {
        info.innerHTML = '<span style="color:var(--danger);">✗ ' + escapeHtml(data.error) + '</span>';
        return;
      }
      // 把读到的内容塞到 importFile，复用现有 UI
      importFile = { name: path.split("/").pop(), headers: data.headers, rows: data.rows };
      document.getElementById("importBody").style.display = "";
      autoMapColumns();
      renderImportMappingArea();
      renderImportPreview();
      updateImportSummary();
      populateDedupKeyOptions();
      info.innerHTML = '<span style="color:var(--success);">✓ 读取成功：' +
        data.headers.length + ' 列 · ' + data.rows.length + ' 行数据</span>';
    })
    .catch(function(e) {
      info.innerHTML = '<span style="color:var(--danger);">✗ ' + escapeHtml(e.message) + '</span>';
    });
}

function buildFileImportWidget() {
  var section = document.createElement("div");
  section.className = "import-section";
  section.innerHTML =
    '<button type="button" class="import-toggle" onclick="toggleImportPanel(this)">' +
      '📊 从 Excel/CSV 批量导入（支持去重）' +
    '</button>' +
    '<div class="import-panel" id="importPanel" style="display:none;">' +
      '<div class="import-config-row">' +
        '<label>导入源</label>' +
        '<div>' +
          '<div style="display:flex;gap:6px;background:#f3f4f6;padding:3px;border-radius:6px;width:fit-content;margin-bottom:8px;">' +
            '<button type="button" id="importSrcBrowserBtn" class="active" onclick="switchImportSource(\'browser\')"' +
                    'style="background:white;color:var(--primary);padding:5px 14px;border:none;border-radius:4px;font-size:12px;cursor:pointer;">' +
              '📂 浏览器选文件' +
            '</button>' +
            '<button type="button" id="importSrcPathBtn" onclick="switchImportSource(\'path\')"' +
                    'style="background:transparent;color:var(--text-2);padding:5px 14px;border:none;border-radius:4px;font-size:12px;cursor:pointer;">' +
              '🗂 服务器文件路径（定时任务可用）' +
            '</button>' +
          '</div>' +
          '<div id="importSrcBrowser">' +
            '<input type="file" id="importFileInput" accept=".xlsx,.xls,.csv" onchange="onImportFileSelected(this)">' +
            '<button type="button" id="importReloadBtn" onclick="reloadImportFile()" ' +
                    'class="ghost small" style="margin-left:8px;display:none;">🔄 重新加载</button>' +
            '<div id="importFileInfo" style="font-size:11px;color:var(--text-3);margin-top:4px;"></div>' +
            '<div style="font-size:11px;color:var(--text-3);margin-top:4px;">' +
              '⚠ 浏览器选的文件是<b>一次性快照</b>，文件后续修改不会自动同步，需要点「🔄 重新加载」。<br>' +
              '⚠ 此模式<b>不能用于定时任务</b>（任务在后台跑时浏览器可能没开）。' +
            '</div>' +
          '</div>' +
          '<div id="importSrcPath" style="display:none;">' +
            '<div style="display:flex;gap:8px;align-items:center;">' +
              '<input type="text" id="importFilePath" placeholder="例：/Users/你的用户名/Downloads/data.csv" ' +
                     'style="flex:1;font-family:ui-monospace,Menlo,monospace;font-size:12px;">' +
              '<button type="button" class="ghost small" onclick="testImportPath()">🔍 测试读取</button>' +
            '</div>' +
            '<div id="importPathInfo" style="font-size:11px;color:var(--text-3);margin-top:4px;"></div>' +
            '<div style="font-size:11px;color:var(--text-3);margin-top:6px;">' +
              '✓ 此模式<b>每次执行都会重新读文件</b>，文件改了不用手动同步。<br>' +
              '✓ 此模式<b>能用于定时任务</b>。<br>' +
              '⚠ 当前后端只支持 <b>.csv</b>（标准库）。.xlsx 需要装 openpyxl（<code>pip install openpyxl</code>）。' +
            '</div>' +
          '</div>' +
        '</div>' +
      '</div>' +
      '<div id="importBody" style="display:none;">' +
        '<div class="import-config-row">' +
          '<label>起始行</label>' +
          '<div>' +
            '<input type="number" id="importStartRow" min="1" value="1" placeholder="默认 1（即第 1 条数据）" oninput="updateImportSummary()" style="width:120px;">' +
            '<span style="font-size:11px;color:var(--text-3);margin-left:8px;">' +
            '行号从「数据行」开始算（不含表头）。1=从第 1 条数据开始。' +
            '</span>' +
          '</div>' +
        '</div>' +
        '<div class="import-config-row">' +
          '<label>去重</label>' +
          '<div>' +
            '<label style="font-size:13px;cursor:pointer;">' +
              '<input type="checkbox" id="importDedup" onchange="onDedupToggle(this)" style="vertical-align:middle;">' +
              ' 启用去重：仅导入 API 表格里 <select id="importDedupKey" disabled style="margin:0 4px;padding:3px 6px;font-size:12px;"></select> 字段不存在的行' +
            '</label>' +
            '<div style="font-size:11px;color:var(--text-3);margin-top:4px;">' +
              '勾选后，导入前会先查 API 表格的所有记录，按所选字段比对，只导入新值。' +
            '</div>' +
          '</div>' +
        '</div>' +
        '<div class="import-config-row" style="grid-template-columns:110px 1fr;align-items:flex-start;">' +
          '<label style="padding-top:6px;">列映射</label>' +
          '<div id="importMappingArea"></div>' +
        '</div>' +
        '<div class="import-config-row" style="grid-template-columns:110px 1fr;align-items:flex-start;">' +
          '<label style="padding-top:6px;">数据预览</label>' +
          '<div class="import-preview" id="importPreview"></div>' +
        '</div>' +
        '<div style="margin-top:14px;">' +
          '<button type="button" id="importExecBtn" onclick="executeBatchImport()" ' +
                  'style="background:var(--primary);color:white;padding:8px 18px;font-size:13px;">' +
            '📥 开始批量导入' +
          '</button>' +
          '<span id="importSummary" style="margin-left:12px;font-size:12px;color:var(--text-3);"></span>' +
        '</div>' +
        '<div id="importProgress" class="import-progress"></div>' +
      '</div>' +
    '</div>';
  return section;
}

function toggleImportPanel(btn) {
  var panel = document.getElementById("importPanel");
  var open = panel.style.display === "none";
  panel.style.display = open ? "" : "none";
  btn.textContent = open
    ? "📊 收起导入面板"
    : "📊 从 Excel/CSV 批量导入（支持去重）";
  if (open && !lastFieldSchema) {
    // 顺便后台拉一下 schema，列映射会用到
    fetchSchemaForFilter();
  }
}

function onImportFileSelected(input) {
  var file = input.files && input.files[0];
  if (!file) return;
  parseImportFile(file);
}

// 用同一个 file 重新解析（用户改了文件内容后用）
function reloadImportFile() {
  var input = document.getElementById("importFileInput");
  var file = input && input.files && input.files[0];
  if (!file) {
    alert("还没选过文件，请先点「Choose File」");
    return;
  }
  parseImportFile(file);
}

function parseImportFile(file) {
  if (window._sheetjsLoadFailed || typeof XLSX === "undefined") {
    alert("Excel 解析库（SheetJS）加载失败，可能是网络问题。\n刷新页面重试，或先把文件另存为 CSV。");
    return;
  }
  var reader = new FileReader();
  reader.onload = function(e) {
    try {
      var data = new Uint8Array(e.target.result);
      var wb = XLSX.read(data, {type: "array"});
      var firstSheet = wb.Sheets[wb.SheetNames[0]];
      // 用 raw:false 让 SheetJS 自动转日期/数字为格式化字符串
      var rows = XLSX.utils.sheet_to_json(firstSheet, {header: 1, defval: "", raw: false});
      if (rows.length === 0) {
        alert("文件是空的或读取失败");
        return;
      }
      var headers = rows[0].map(function(h) { return String(h || "").trim(); });
      var dataRows = rows.slice(1);
      // 过滤掉完全空的尾部行（Excel 末尾常有看不见的空行）
      while (dataRows.length > 0) {
        var last = dataRows[dataRows.length - 1];
        var allEmpty = (last || []).every(function(c) { return c == null || String(c).trim() === ""; });
        if (allEmpty) dataRows.pop();
        else break;
      }
      importFile = { name: file.name, headers: headers, rows: dataRows };
      document.getElementById("importBody").style.display = "";
      document.getElementById("importReloadBtn").style.display = "";
      var info = document.getElementById("importFileInfo");
      if (info) {
        info.innerHTML = '<span style="color:var(--success);">✓ 已解析「' +
          escapeHtml(file.name) + '」</span> · 表头 ' + headers.length +
          ' 列 · 数据 <b style="color:var(--text);">' + dataRows.length + '</b> 行 · ' +
          '于 ' + new Date().toLocaleTimeString() + ' 加载';
      }
      autoMapColumns();
      renderImportMappingArea();
      renderImportPreview();
      updateImportSummary();
      populateDedupKeyOptions();
    } catch (err) {
      alert("解析文件失败：" + err.message);
    }
  };
  reader.readAsArrayBuffer(file);
}

// 自动按字段名匹配文件列 → AITable 字段
function autoMapColumns() {
  importMapping = {};
  if (!importFile) return;
  var aitableNames = (lastFieldSchema || []).map(function(f) { return f.name; });
  importFile.headers.forEach(function(h, idx) {
    if (aitableNames.indexOf(h) >= 0) {
      importMapping[idx] = h;
    } else {
      importMapping[idx] = "";  // 默认忽略
    }
  });
}

function renderImportMappingArea() {
  var area = document.getElementById("importMappingArea");
  if (!importFile) { area.innerHTML = ""; return; }
  var aitableNames = (lastFieldSchema || []).map(function(f) { return f.name; });
  var html = '<table class="import-mapping-table">';
  html += '<thead><tr><th style="width:40%;">文件列（前 1 行预览）</th><th class="arrow">→</th><th>AITable 字段</th></tr></thead><tbody>';
  importFile.headers.forEach(function(h, idx) {
    var sample = importFile.rows.length > 0 ? String(importFile.rows[0][idx] || "") : "";
    if (sample.length > 30) sample = sample.substring(0, 30) + "...";
    html += '<tr>';
    html += '<td><span class="col-name">' + escapeHtml(h) + '</span>' +
            (sample ? '<span class="preview-val">e.g. ' + escapeHtml(sample) + '</span>' : '') + '</td>';
    html += '<td class="arrow">→</td>';
    html += '<td><select onchange="importMapping[' + idx + ']=this.value;updateImportSummary();">';
    html += '<option value="">(忽略此列)</option>';
    aitableNames.forEach(function(name) {
      html += '<option value="' + escapeHtml(name) + '"' +
              (importMapping[idx] === name ? " selected" : "") + '>' + escapeHtml(name) + '</option>';
    });
    html += '</select></td>';
    html += '</tr>';
  });
  html += '</tbody></table>';
  if (!lastFieldSchema) {
    html += '<div style="margin-top:6px;font-size:11px;color:var(--text-3);">⚠ 还没拉到 AITable 字段列表，下拉是空的。请先填好表格 ID，或点上方「📋 查询整张表」一次。</div>';
  }
  area.innerHTML = html;
}

function renderImportPreview() {
  var preview = document.getElementById("importPreview");
  if (!importFile) { preview.innerHTML = ""; return; }
  var html = '<table><thead><tr>';
  importFile.headers.forEach(function(h) { html += '<th>' + escapeHtml(h) + '</th>'; });
  html += '</tr></thead><tbody>';
  importFile.rows.slice(0, 5).forEach(function(row) {
    html += '<tr>';
    importFile.headers.forEach(function(_, idx) {
      var v = row[idx];
      var t = String(v == null ? "" : v);
      if (t.length > 50) t = t.substring(0, 50) + "...";
      html += '<td>' + escapeHtml(t) + '</td>';
    });
    html += '</tr>';
  });
  html += '</tbody></table>';
  if (importFile.rows.length > 5) {
    html += '<div style="padding:6px 10px;font-size:11px;color:var(--text-3);background:#f8fafc;">... 共 ' + importFile.rows.length + ' 行数据，仅显示前 5 行</div>';
  }
  preview.innerHTML = html;
}

function populateDedupKeyOptions() {
  var sel = document.getElementById("importDedupKey");
  sel.innerHTML = "";
  // 从已映射的字段里选
  var mapped = Object.values(importMapping).filter(function(v) { return v; });
  if (mapped.length === 0) {
    sel.innerHTML = '<option value="">(没有已映射的字段)</option>';
    return;
  }
  // 去重
  var unique = [];
  mapped.forEach(function(v) { if (unique.indexOf(v) < 0) unique.push(v); });
  unique.forEach(function(name) {
    var o = document.createElement("option");
    o.value = name; o.textContent = name;
    sel.appendChild(o);
  });
}

function onDedupToggle(cb) {
  document.getElementById("importDedupKey").disabled = !cb.checked;
  if (cb.checked) populateDedupKeyOptions();
}

function updateImportSummary() {
  var s = document.getElementById("importSummary");
  if (!importFile) { s.textContent = ""; return; }
  var mapped = Object.values(importMapping).filter(function(v) { return v; });
  var startRow = parseInt(document.getElementById("importStartRow").value, 10) || 1;
  var totalAfterStart = Math.max(0, importFile.rows.length - (startRow - 1));
  s.textContent = "文件共 " + importFile.rows.length + " 行 · 从第 " + startRow + " 行起 = " +
                  totalAfterStart + " 条待导入 · 已映射 " + mapped.length + " 列";
  populateDedupKeyOptions();
}

// 简单的类型转换：参考 schema 的 type 把字符串转成合适的类型
function coerceForFieldType(value, fieldName) {
  if (!lastFieldSchema) return value;
  var schemaField = null;
  for (var i = 0; i < lastFieldSchema.length; i++) {
    if (lastFieldSchema[i].name === fieldName) { schemaField = lastFieldSchema[i]; break; }
  }
  if (!schemaField) return value;
  var t = schemaField.type;
  if (value == null || value === "") return null;
  var s = String(value).trim();
  if (t === "Number" || t === "Currency" || t === "Percent" || t === "Rating" || t === "AutoNumber") {
    var n = parseFloat(s);
    return isNaN(n) ? value : n;
  }
  if (t === "Checkbox") {
    var lower = s.toLowerCase();
    if (lower === "true" || lower === "1" || lower === "是" || lower === "yes" || lower === "✓") return true;
    if (lower === "false" || lower === "0" || lower === "否" || lower === "no" || lower === "") return false;
    return Boolean(s);
  }
  if (t === "DateTime") {
    // 尝试解析日期
    var d = new Date(s);
    if (!isNaN(d.getTime())) return d.getTime();
    return value;
  }
  if (t === "MultiSelect" && typeof s === "string") {
    return s.split(/[,，;；]/).map(function(x) { return x.trim(); }).filter(Boolean);
  }
  return value;
}

function executeBatchImport() {
  if (!importFile) { alert("请先选择文件"); return; }

  var token = document.getElementById("token").value.trim();
  var baseUrl = document.getElementById("baseUrl").value.trim();
  var dstInput = document.getElementById("param_datasheetId");
  var datasheetId = dstInput ? dstInput.value.trim() : "";
  if (!token || !baseUrl || !datasheetId) {
    alert("请先填好 顶部 Token / 接口地址 / 表格 ID");
    return;
  }

  var startRow = parseInt(document.getElementById("importStartRow").value, 10) || 1;
  var dedup = document.getElementById("importDedup").checked;
  var dedupKey = dedup ? document.getElementById("importDedupKey").value : "";
  if (dedup && !dedupKey) { alert("勾选了去重但没选去重字段"); return; }

  var mapped = Object.values(importMapping).filter(function(v) { return v; });
  if (mapped.length === 0) { alert("还没映射任何列"); return; }

  // 切片：从 startRow 开始（startRow=1 表示第 1 条数据）
  var dataRows = importFile.rows.slice(startRow - 1);
  if (dataRows.length === 0) { alert("起始行后面没有数据"); return; }

  // 转换为 AITable records
  var records = dataRows.map(function(row) {
    var fields = {};
    Object.keys(importMapping).forEach(function(idxStr) {
      var idx = parseInt(idxStr, 10);
      var aitableField = importMapping[idx];
      if (!aitableField) return;
      var raw = row[idx];
      if (raw == null || raw === "") return;
      fields[aitableField] = coerceForFieldType(raw, aitableField);
    });
    return { fields: fields };
  }).filter(function(r) {
    return Object.keys(r.fields).length > 0;  // 跳过完全空的行
  });

  if (records.length === 0) { alert("处理后没有有效数据"); return; }

  var progressEl = document.getElementById("importProgress");
  progressEl.classList.add("active");
  progressEl.innerHTML = "📂 文件 " + importFile.name + "<br>📊 待导入 " + records.length + " 条<br>";

  var execBtn = document.getElementById("importExecBtn");
  execBtn.disabled = true;
  execBtn.textContent = "⏳ 导入中...";

  // 步骤 1：去重（如果勾选）
  var prep;
  if (dedup) {
    progressEl.innerHTML += "<br>🔍 拉取 API 表格的现有记录用于去重...<br>";
    prep = fetchAllExistingValues(token, baseUrl, datasheetId, dedupKey).then(function(info) {
      progressEl.innerHTML +=
        "  → API 表格共 " + info.totalRecords + " 条记录，其中 " +
        info.recordsWithValue + " 条「" + dedupKey + "」字段非空，去重后 " +
        info.uniqueValues.size + " 个唯一值<br>";
      var existingSet = info.uniqueValues;
      var before = records.length;
      var filtered = records.filter(function(r) {
        var v = r.fields[dedupKey];
        if (v == null) return true;
        return !existingSet.has(String(v));
      });
      progressEl.innerHTML += "  → 文件 " + before + " 条 → 过滤后剩 <b>" + filtered.length +
                              "</b> 条新增（跳过 " + (before - filtered.length) + " 条已存在）<br>";
      return filtered;
    });
  } else {
    prep = Promise.resolve(records);
  }

  prep.then(function(toImport) {
    if (toImport.length === 0) {
      progressEl.innerHTML += "<br><span class='ok'>✓ 没有需要新增的记录，导入结束</span>";
      execBtn.disabled = false; execBtn.textContent = "📥 开始批量导入";
      return;
    }
    return doBatchedCreate(toImport, token, baseUrl, datasheetId, progressEl).then(function() {
      execBtn.disabled = false; execBtn.textContent = "📥 开始批量导入";
    });
  }).catch(function(e) {
    progressEl.innerHTML += "<br><span class='err'>✗ " + escapeHtml(e.message || String(e)) + "</span>";
    execBtn.disabled = false; execBtn.textContent = "📥 开始批量导入";
  });
}

// 拉所有现有记录某个字段的值（处理分页）
// 返回 { uniqueValues: Set, totalRecords: number, recordsWithValue: number }
function fetchAllExistingValues(token, baseUrl, datasheetId, fieldName) {
  var existing = new Set();
  var totalRecords = 0;
  var recordsWithValue = 0;
  var page = 1, pageSize = 1000;
  function fetchPage() {
    return fetch("/api/call", {
      method: "POST", headers: {"Content-Type": "application/json"},
      body: JSON.stringify({
        op: "get_records", token: token, baseUrl: baseUrl, datasheetId: datasheetId,
        params: {
          datasheetId: datasheetId,
          fieldKey: "name",
          fields: fieldName,
          pageSize: String(pageSize),
          pageNum: String(page)
        }
      })
    }).then(function(r) { return r.json(); })
      .then(function(data) {
        if (data.status >= 400) {
          throw new Error("查询现有记录失败 HTTP " + data.status);
        }
        var records = extractRecords(data.response) || [];
        totalRecords += records.length;
        records.forEach(function(r) {
          var v = r.fields ? r.fields[fieldName] : null;
          if (v != null && v !== "") {
            existing.add(String(v));
            recordsWithValue++;
          }
        });
        if (records.length < pageSize) {
          return { uniqueValues: existing, totalRecords: totalRecords, recordsWithValue: recordsWithValue };
        }
        page++;
        return fetchPage();
      });
  }
  return fetchPage();
}

// 把 records 切成 10 条/批，串行 POST
function doBatchedCreate(records, token, baseUrl, datasheetId, progressEl) {
  var BATCH = 10;
  var batches = [];
  for (var i = 0; i < records.length; i += BATCH) {
    batches.push(records.slice(i, i + BATCH));
  }
  progressEl.innerHTML += "<br>📦 分 " + batches.length + " 批发送（每批最多 " + BATCH + " 条）...<br>";

  var success = 0, fail = 0;
  var idx = 0;
  function next() {
    if (idx >= batches.length) {
      progressEl.innerHTML += "<br><span class='ok'>✓ 全部完成：成功 " + success + " 条，失败 " + fail + " 条</span>";
      return;
    }
    var batch = batches[idx];
    var batchNo = idx + 1;
    progressEl.innerHTML += "  → 第 " + batchNo + "/" + batches.length + " 批 (" + batch.length + " 条) ... ";
    idx++;
    return fetch("/api/call", {
      method: "POST", headers: {"Content-Type": "application/json"},
      body: JSON.stringify({
        op: "create", token: token, baseUrl: baseUrl, datasheetId: datasheetId,
        params: {
          datasheetId: datasheetId,
          fieldKey: "name",
          records: JSON.stringify(batch)
        }
      })
    }).then(function(r) { return r.json(); })
      .then(function(data) {
        if (data.status >= 200 && data.status < 300) {
          success += batch.length;
          progressEl.innerHTML += "<span class='ok'>✓ HTTP " + data.status + "</span><br>";
        } else {
          fail += batch.length;
          var err = "";
          try { err = (data.response && data.response.message) || JSON.stringify(data.response).substring(0, 100); }
          catch (e) { err = String(data.response).substring(0, 100); }
          progressEl.innerHTML += "<span class='err'>✗ HTTP " + data.status + " " + escapeHtml(err) + "</span><br>";
        }
        progressEl.scrollTop = progressEl.scrollHeight;
        return next();
      })
      .catch(function(e) {
        fail += batch.length;
        progressEl.innerHTML += "<span class='err'>✗ " + escapeHtml(e.message) + "</span><br>";
        return next();
      });
  }
  return next();
}

// 在新标签页打开 APITable 的表格页
function openInAitable() {
  if (!lastDatasheetId) { flash("没有表格 ID"); return; }
  // 把 baseUrl 末尾的 /fusion/v1 去掉，得到 origin
  var origin = (lastBaseUrl || document.getElementById("baseUrl").value)
                 .replace(/\/fusion\/v1\/?$/, "");
  var url = origin + "/workbench/" + lastDatasheetId;
  window.open(url, "_blank");
}

// ---------- 基础配置 ----------
function saveBaseConfig() {
  var cfg = {
    baseUrl: document.getElementById("baseUrl").value,
    token: document.getElementById("token").value
  };
  localStorage.setItem("aitable_baseconfig", JSON.stringify(cfg));
  flash("基础配置已保存");
}

function loadBaseConfig() {
  var raw = localStorage.getItem("aitable_baseconfig");
  if (!raw) { flash("暂无保存的基础配置"); return; }
  var cfg = JSON.parse(raw);
  if (cfg.baseUrl) document.getElementById("baseUrl").value = cfg.baseUrl;
  if (cfg.token) document.getElementById("token").value = cfg.token;
  flash("已加载");
}

function flash(msg) {
  var el = document.getElementById("result");
  el.textContent = "ℹ " + msg;
  el.className = "result-box";
  setTimeout(function() {
    if (el.textContent === "ℹ " + msg) { el.textContent = ""; }
  }, 1800);
}

// ---------- 模板管理 ----------
function loadTemplatesList() {
  var raw = localStorage.getItem("aitable_templates");
  return raw ? JSON.parse(raw) : [];
}
function saveTemplatesList(arr) { localStorage.setItem("aitable_templates", JSON.stringify(arr)); }

function refreshTemplateSelect() {
  var sel = document.getElementById("templateSelect");
  var prev = sel.value;
  var tpls = loadTemplatesList();
  sel.innerHTML = '<option value="">— 选择已保存的模板 —</option>';
  tpls.forEach(function(t, i) {
    var o = document.createElement("option");
    o.value = String(i);
    o.textContent = t.name + "（" + opLabel(t.op) + "）";
    sel.appendChild(o);
  });
  if (prev && sel.querySelector('option[value="' + prev + '"]')) sel.value = prev;
}

function opLabel(op) {
  var map = {
    get_records: "查询记录", create: "创建记录", update: "更新记录",
    delete: "删除记录", get_fields: "字段列表", get_views: "视图列表",
    get_spaces: "空间站列表"
  };
  return map[op] || op;
}

function collectCurrentState() {
  var op = document.getElementById("operation").value;
  var formFields = FORMS[op] || [];
  var paramValues = {};
  formFields.forEach(function(f) {
    if (f.type === "records") {
      paramValues[f.key + "__visual"] = recordsState[f.mode];
      paramValues[f.key + "__jsonMode"] = jsonModeState[f.mode];
      if (jsonModeState[f.mode]) {
        var ta = document.getElementById("param_records_json_" + f.mode);
        paramValues[f.key + "__json"] = ta ? ta.value : "";
      }
      // 文件导入配置（不保存文件本身）
      if (f.showFileImport) {
        var startRow = document.getElementById("importStartRow");
        var dedup = document.getElementById("importDedup");
        var dedupKey = document.getElementById("importDedupKey");
        if (startRow) paramValues["__importStartRow"] = startRow.value;
        if (dedup) paramValues["__importDedup"] = dedup.checked;
        if (dedupKey) paramValues["__importDedupKey"] = dedupKey.value;
        paramValues["__importMapping"] = JSON.stringify(importMapping);
      }
    } else {
      var el = document.getElementById("param_" + f.key);
      if (el) paramValues[f.key] = el.value;
    }
  });
  return { op: op, params: paramValues };
}

function applyState(state) {
  document.getElementById("operation").value = state.op;
  renderParams();
  var formFields = FORMS[state.op] || [];
  formFields.forEach(function(f) {
    if (f.type === "records") {
      var visual = state.params[f.key + "__visual"];
      if (visual) recordsState[f.mode] = visual;
      jsonModeState[f.mode] = !!state.params[f.key + "__jsonMode"];
      var ed = document.getElementById("records_editor_" + f.mode);
      if (ed) {
        renderRecordsEditor(ed, f.mode);
        if (jsonModeState[f.mode] && state.params[f.key + "__json"]) {
          var ta = document.getElementById("param_records_json_" + f.mode);
          if (ta) ta.value = state.params[f.key + "__json"];
        }
      }
      // 恢复文件导入配置（注意：模板不存文件本身，需要重新选）
      if (f.showFileImport) {
        if (state.params["__importMapping"]) {
          try { importMapping = JSON.parse(state.params["__importMapping"]) || {}; }
          catch (e) {}
        }
        // 配置 input 的值（如果导入面板没打开，下次打开时已存好的值会出现）
        setTimeout(function() {
          var sr = document.getElementById("importStartRow");
          var dd = document.getElementById("importDedup");
          var dk = document.getElementById("importDedupKey");
          if (sr && state.params["__importStartRow"] != null) sr.value = state.params["__importStartRow"];
          if (dd) {
            dd.checked = !!state.params["__importDedup"];
            if (dk) dk.disabled = !dd.checked;
          }
          if (dk && state.params["__importDedupKey"]) dk.value = state.params["__importDedupKey"];
        }, 0);
      }
    } else {
      var el = document.getElementById("param_" + f.key);
      if (el && state.params[f.key] != null) el.value = state.params[f.key];
    }
  });
}

function countNonEmpty(state) {
  // 统计实际填了内容的参数（默认值 fieldKey="name" 也算填了）
  var n = 0;
  Object.keys(state.params).forEach(function(k) {
    if (k.indexOf("__") >= 0) {
      // 记录编辑器：visual 是数组，看是否有非空字段
      if (k.endsWith("__visual")) {
        var arr = state.params[k];
        if (arr && arr.length > 0) {
          arr.forEach(function(rec) {
            if (rec.recordId) n++;
            (rec.fields || []).forEach(function(fld) {
              if (fld.name || fld.value) n++;
            });
          });
        }
      } else if (k.endsWith("__json") && state.params[k]) {
        n++;
      }
      return;
    }
    var v = state.params[k];
    if (v !== "" && v != null) n++;
  });
  return n;
}

function saveAsTemplate() {
  var state = collectCurrentState();
  var n = countNonEmpty(state);
  if (n === 0) {
    if (!confirm("当前所有参数都是空的，确定要保存空模板吗？")) return;
  }
  var name = prompt("起个名字：");
  if (!name) return;
  var tpls = loadTemplatesList();
  tpls.push({ name: name, op: state.op, params: state.params });
  saveTemplatesList(tpls);
  refreshTemplateSelect();
  document.getElementById("templateSelect").value = String(tpls.length - 1);
  flash("已保存模板「" + name + "」（含 " + n + " 个非空参数）");
}

function overwriteTemplate() {
  var sel = document.getElementById("templateSelect");
  var idx = sel.value;
  if (idx === "") { flash("请先在下拉框选择一个要覆盖的模板"); return; }
  var tpls = loadTemplatesList();
  var t = tpls[parseInt(idx, 10)];
  if (!t) return;
  if (!confirm('确定用当前内容覆盖模板「' + t.name + '」？')) return;
  var state = collectCurrentState();
  t.op = state.op; t.params = state.params;
  saveTemplatesList(tpls);
  flash("已覆盖：" + t.name);
}

function loadTemplate() {
  var sel = document.getElementById("templateSelect");
  var idx = sel.value;
  if (idx === "") { flash("请先选择一个模板"); return; }
  var tpls = loadTemplatesList();
  var t = tpls[parseInt(idx, 10)];
  if (!t) return;
  applyState(t);

  // 反馈：把模板里实际包含的参数详细列出来
  var formFields = FORMS[t.op] || [];
  var lines = ["✓ 已加载模板「" + t.name + "」",
               "  操作类型：" + opLabel(t.op),
               "──────────────────────────────────────────────",
               "模板里包含的参数："];
  var found = 0;
  formFields.forEach(function(f) {
    if (f.type === "records") {
      var visual = t.params[f.key + "__visual"];
      if (visual && visual.length > 0) {
        lines.push("  • 记录数据（" + visual.length + " 条记录）");
        found++;
      }
    } else {
      var v = t.params[f.key];
      if (v !== "" && v != null) {
        lines.push("  • " + f.label + " = " + v);
        found++;
      }
    }
  });
  if (found === 0) {
    lines.push("  ⚠ 这个模板是空的，所有参数都没填");
    lines.push("");
    lines.push("提示：先在参数区填好数据，再点「覆盖当前模板」即可保存进去");
  }
  showResult(lines.join("\n"), found === 0);

  // 高亮参数卡片
  var paramsCard = document.getElementById("params").closest(".card");
  if (paramsCard) {
    paramsCard.style.transition = "box-shadow 0.4s, border-color 0.4s";
    paramsCard.style.borderColor = "var(--primary)";
    paramsCard.style.boxShadow = "0 0 0 3px rgba(99,102,241,0.18)";
    setTimeout(function() {
      paramsCard.style.borderColor = "";
      paramsCard.style.boxShadow = "";
    }, 1200);
  }
}

function deleteTemplate() {
  var sel = document.getElementById("templateSelect");
  var idx = sel.value;
  if (idx === "") { flash("请先选择要删除的模板"); return; }
  var tpls = loadTemplatesList();
  var t = tpls[parseInt(idx, 10)];
  if (!t) return;
  if (!confirm('确定删除模板「' + t.name + '」？')) return;
  tpls.splice(parseInt(idx, 10), 1);
  saveTemplatesList(tpls);
  refreshTemplateSelect();
  flash("已删除");
}

// ---------- 定时任务 ----------
function addScheduleFromCurrent() {
  var op = document.getElementById("operation").value;
  var token = document.getElementById("token").value.trim();
  var baseUrl = document.getElementById("baseUrl").value.trim();
  if (!token) { alert("请先填写 API 密钥"); return; }
  if (!baseUrl) { alert("请先填写接口地址"); return; }

  var formFields = FORMS[op] || [];
  var params = {};
  for (var i = 0; i < formFields.length; i++) {
    var f = formFields[i];
    params[f.key] = getParamValue(f.key, op);
  }
  if (op !== "get_spaces" && !params.datasheetId) {
    alert("请先填写表格 ID"); return;
  }

  // 检查文件导入：浏览器选的文件不能用于定时任务，必须用服务器路径
  var importConfig = null;
  if (op === "create" && importFile) {
    if (importSource === "browser") {
      var hint = "⚠ 你当前是「浏览器选文件」模式，浏览器选的文件无法在后台被读到，定时任务跑不了。\n\n" +
                 "需要切到「🗂 服务器文件路径」模式，填一个本地文件路径（如 /Users/.../data.csv），" +
                 "后端每次执行时会重新读这个文件。\n\n" +
                 "现在切到路径模式吗？";
      if (confirm(hint)) {
        switchImportSource("path");
      }
      return;
    }
    // path 模式
    var filePath = (document.getElementById("importFilePath") || {}).value || "";
    filePath = filePath.trim();
    if (!filePath) {
      alert("请在导入面板里填好「服务器文件路径」，并点「测试读取」确认能读到");
      return;
    }
    importConfig = {
      filePath: filePath,
      mapping: importMapping,
      startRow: parseInt((document.getElementById("importStartRow") || {}).value, 10) || 1,
      dedup: (document.getElementById("importDedup") || {}).checked || false,
      dedupKey: (document.getElementById("importDedupKey") || {}).value || ""
    };
  }

  // 弹模态对话框选调度方式
  showScheduleDialog(op, params, importConfig, token, baseUrl);
}

function showScheduleDialog(op, params, importConfig, token, baseUrl) {
  // 移除已有对话框
  var existing = document.getElementById("schedDialog");
  if (existing) existing.remove();

  var defaultName = opLabel(op);
  if (params.datasheetId) defaultName += " - " + params.datasheetId.substring(0, 12);
  if (importConfig) defaultName = "导入 " + (importConfig.filePath.split("/").pop() || "");

  var modal = document.createElement("div");
  modal.id = "schedDialog";
  modal.style.cssText = "position:fixed;inset:0;background:rgba(15,23,42,0.5);display:flex;align-items:center;justify-content:center;z-index:9999;";
  modal.innerHTML =
    '<div style="background:white;border-radius:12px;padding:24px;max-width:520px;width:90%;box-shadow:0 20px 60px rgba(0,0,0,0.3);">' +
    '<h3 style="margin:0 0 16px;font-size:18px;">🔁 添加定时任务</h3>' +
    '<div style="margin-bottom:14px;">' +
      '<label style="display:block;font-size:12px;color:var(--text-2);margin-bottom:4px;">名称</label>' +
      '<input type="text" id="schedName" value="' + defaultName + '" style="width:100%;padding:8px 10px;border:1px solid var(--border-strong);border-radius:6px;">' +
    '</div>' +
    '<div style="margin-bottom:14px;">' +
      '<label style="display:block;font-size:12px;color:var(--text-2);margin-bottom:6px;">调度方式</label>' +
      '<div style="display:flex;gap:6px;background:#f3f4f6;padding:3px;border-radius:6px;width:fit-content;">' +
        '<button type="button" id="schedModeInterval" class="active" onclick="switchSchedMode(\'interval\')" ' +
                'style="background:white;color:var(--primary);padding:6px 14px;border:none;border-radius:4px;font-size:12px;cursor:pointer;">⏱ 每隔 N 分钟</button>' +
        '<button type="button" id="schedModeWeekly" onclick="switchSchedMode(\'weekly\')" ' +
                'style="background:transparent;color:var(--text-2);padding:6px 14px;border:none;border-radius:4px;font-size:12px;cursor:pointer;">📅 每周固定时间</button>' +
      '</div>' +
    '</div>' +
    '<div id="schedConfigInterval" style="margin-bottom:14px;">' +
      '<label style="display:block;font-size:12px;color:var(--text-2);margin-bottom:4px;">间隔（分钟）</label>' +
      '<input type="number" id="schedInterval" min="1" value="60" style="width:120px;padding:8px 10px;border:1px solid var(--border-strong);border-radius:6px;">' +
      '<div style="font-size:11px;color:var(--text-3);margin-top:4px;">常用：5、15、30、60、360（6h）、720（12h）、1440（每天）</div>' +
    '</div>' +
    '<div id="schedConfigWeekly" style="display:none;margin-bottom:14px;">' +
      '<label style="display:block;font-size:12px;color:var(--text-2);margin-bottom:6px;">星期几（可多选）</label>' +
      '<div id="schedDaysOfWeek" style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:12px;">' +
        ['一','二','三','四','五','六','日'].map(function(d, i) {
          return '<label class="sched-day" data-day="' + (i+1) + '" ' +
                 'style="cursor:pointer;padding:8px 14px;border:1px solid var(--border-strong);border-radius:6px;font-size:13px;user-select:none;">' +
                 '<input type="checkbox" style="display:none;" onchange="toggleSchedDay(this)">周' + d +
                 '</label>';
        }).join('') +
      '</div>' +
      '<label style="display:block;font-size:12px;color:var(--text-2);margin-bottom:4px;">时间</label>' +
      '<input type="time" id="schedTime" value="09:00" style="padding:8px 10px;border:1px solid var(--border-strong);border-radius:6px;font-size:14px;">' +
      '<div style="font-size:11px;color:var(--text-3);margin-top:4px;">每个选中的星期几，到达此时间后触发一次</div>' +
    '</div>' +
    (importConfig
      ? ('<div style="background:var(--primary-soft);padding:10px 12px;border-radius:6px;margin-bottom:14px;font-size:12px;color:var(--primary);">' +
         '📥 文件导入任务：将定时读 <code>' + escapeHtml(importConfig.filePath) + '</code>，' +
         '映射 ' + Object.values(importConfig.mapping).filter(function(v){return v;}).length + ' 列' +
         (importConfig.dedup ? '，按「' + escapeHtml(importConfig.dedupKey) + '」去重' : '') +
         '</div>')
      : '') +
    '<div style="display:flex;gap:8px;justify-content:flex-end;margin-top:18px;">' +
      '<button type="button" onclick="document.getElementById(\'schedDialog\').remove()" ' +
              'style="padding:8px 16px;background:white;color:var(--text);border:1px solid var(--border-strong);border-radius:6px;cursor:pointer;">取消</button>' +
      '<button type="button" id="schedConfirmBtn" ' +
              'style="padding:8px 18px;background:var(--primary);color:white;border:none;border-radius:6px;cursor:pointer;">添加</button>' +
    '</div>' +
    '</div>';
  document.body.appendChild(modal);

  // 状态
  modal._mode = "interval";
  modal._daysOfWeek = [];
  modal._op = op;
  modal._params = params;
  modal._importConfig = importConfig;
  modal._token = token;
  modal._baseUrl = baseUrl;

  document.getElementById("schedConfirmBtn").onclick = submitSchedule;
}

function switchSchedMode(mode) {
  var modal = document.getElementById("schedDialog");
  if (!modal) return;
  modal._mode = mode;
  document.getElementById("schedModeInterval").style.background = mode === "interval" ? "white" : "transparent";
  document.getElementById("schedModeInterval").style.color = mode === "interval" ? "var(--primary)" : "var(--text-2)";
  document.getElementById("schedModeWeekly").style.background = mode === "weekly" ? "white" : "transparent";
  document.getElementById("schedModeWeekly").style.color = mode === "weekly" ? "var(--primary)" : "var(--text-2)";
  document.getElementById("schedConfigInterval").style.display = mode === "interval" ? "" : "none";
  document.getElementById("schedConfigWeekly").style.display = mode === "weekly" ? "" : "none";
}

function toggleSchedDay(cb) {
  var modal = document.getElementById("schedDialog");
  var label = cb.parentElement;
  var day = parseInt(label.dataset.day, 10);
  if (cb.checked) {
    if (modal._daysOfWeek.indexOf(day) < 0) modal._daysOfWeek.push(day);
    label.style.background = "var(--primary)";
    label.style.color = "white";
    label.style.borderColor = "var(--primary)";
  } else {
    modal._daysOfWeek = modal._daysOfWeek.filter(function(d) { return d !== day; });
    label.style.background = "white";
    label.style.color = "var(--text)";
    label.style.borderColor = "var(--border-strong)";
  }
}

function submitSchedule() {
  var modal = document.getElementById("schedDialog");
  if (!modal) return;
  var name = document.getElementById("schedName").value.trim();
  if (!name) { alert("请填写名称"); return; }

  var body = {
    name: name,
    enabled: true,
    scheduleType: modal._mode,
    op: modal._op,
    token: modal._token,
    baseUrl: modal._baseUrl,
    datasheetId: modal._params.datasheetId || "",
    params: modal._params,
    importConfig: modal._importConfig || null
  };

  if (modal._mode === "interval") {
    var iv = parseInt(document.getElementById("schedInterval").value, 10);
    if (isNaN(iv) || iv < 1) { alert("间隔必须 >= 1"); return; }
    body.intervalMinutes = iv;
  } else {
    if (modal._daysOfWeek.length === 0) { alert("请至少选一个星期几"); return; }
    body.daysOfWeek = modal._daysOfWeek;
    body.time = document.getElementById("schedTime").value || "09:00";
  }

  fetch("/api/schedules", {
    method: "POST", headers: {"Content-Type": "application/json"},
    body: JSON.stringify(body)
  }).then(function(r) { return r.json(); })
    .then(function() {
      var desc;
      if (modal._mode === "interval") desc = "每 " + body.intervalMinutes + " 分钟";
      else {
        var dayNames = ['','一','二','三','四','五','六','日'];
        desc = "周" + body.daysOfWeek.sort().map(function(d){ return dayNames[d]; }).join("、") + " " + body.time;
      }
      flash("✓ 已添加定时任务「" + name + "」（" + desc + "）");
      modal.remove();
      refreshSchedules();
    })
    .catch(function(e) { alert("添加失败: " + e.message); });
}

function refreshSchedules() {
  fetch("/api/schedules").then(function(r) { return r.json(); })
    .then(function(schedules) {
      var container = document.getElementById("schedulesList");
      if (!schedules || schedules.length === 0) {
        container.innerHTML = '<div class="empty-state">还没有定时任务<br><span style="font-size:12px;">在上面填好参数后，点「＋ 把当前操作设为定时任务」</span></div>';
        return;
      }

      var html = '<table class="schedules-table"><thead><tr>';
      html += '<th>启用</th><th>名称</th><th>操作</th><th>调度</th><th>上次运行</th><th></th>';
      html += '</tr></thead><tbody>';
      var dayNames = ['','一','二','三','四','五','六','日'];
      schedules.forEach(function(s) {
        var lastRun = s.lastRun
          ? formatTimestamp(s.lastRun) + " · HTTP " + (s.lastStatus || "?")
          : "尚未运行";
        var statusClass = "status-pending";
        if (s.lastStatus && s.lastStatus >= 200 && s.lastStatus < 300) statusClass = "status-ok";
        else if (s.lastStatus && s.lastStatus >= 400) statusClass = "status-err";

        var schedDesc;
        if ((s.scheduleType || "interval") === "weekly") {
          var ds = (s.daysOfWeek || []).slice().sort();
          schedDesc = "📅 周" + ds.map(function(d) { return dayNames[d] || d; }).join("、") +
                      " " + (s.time || "09:00");
        } else {
          schedDesc = "⏱ 每 " + (s.intervalMinutes || 60) + " 分钟";
        }
        if (s.importConfig) schedDesc += " · 📥 文件导入";

        html += '<tr>';
        html += '<td><label class="switch"><input type="checkbox"' + (s.enabled ? " checked" : "")
              + ' onchange="toggleSchedule(\'' + s.id + '\', this.checked)"><span class="slider"></span></label></td>';
        html += '<td class="name-col">' + escapeHtml(s.name) + '</td>';
        html += '<td>' + escapeHtml(opLabel(s.op)) + '</td>';
        html += '<td>' + escapeHtml(schedDesc) + '</td>';
        html += '<td class="' + statusClass + '">' + escapeHtml(lastRun) + '</td>';
        html += '<td><div class="row-actions">';
        html += '<button class="ghost small" onclick="runScheduleNow(\'' + s.id + '\')">立即运行</button>';
        html += '<button class="ghost small" style="color:var(--danger);" onclick="deleteSchedule(\'' + s.id + '\')">删除</button>';
        html += '</div></td>';
        html += '</tr>';
      });
      html += '</tbody></table>';
      container.innerHTML = html;
    })
    .catch(function(e) {
      document.getElementById("schedulesList").innerHTML =
        '<div class="empty-state status-err">加载失败: ' + escapeHtml(e.message) + '</div>';
    });
}

function toggleSchedule(id, enabled) {
  fetch("/api/schedules/" + id, {
    method: "PATCH",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify({enabled: enabled})
  }).then(function() {
    flash(enabled ? "已启用" : "已暂停");
    refreshSchedules();
  });
}

function deleteSchedule(id) {
  if (!confirm("确定删除该定时任务？")) return;
  fetch("/api/schedules/" + id, {method: "DELETE"})
    .then(function() { flash("已删除"); refreshSchedules(); });
}

function runScheduleNow(id) {
  flash("执行中...");
  fetch("/api/schedules/run/" + id, {method: "POST"})
    .then(function(r) { return r.json(); })
    .then(function(data) {
      var ok = data.status >= 200 && data.status < 300;
      var text = (ok ? "✓" : "✗") + " 执行完成 · HTTP " + data.status + "\n";
      try { text += JSON.stringify(JSON.parse(data.response), null, 2); }
      catch (e) { text += data.response || ""; }
      showResult(text, !ok);
      refreshSchedules();
      document.getElementById("result").scrollIntoView({behavior: "smooth", block: "center"});
    })
    .catch(function(e) { flash("执行失败: " + e.message); });
}

function showScheduleLogs() {
  fetch("/api/schedule-logs").then(function(r) { return r.json(); })
    .then(function(logs) {
      if (!logs || logs.length === 0) {
        showResult("暂无执行日志", false); return;
      }
      var lines = ["📜 最近 " + logs.length + " 条执行日志（按时间倒序）",
                   "──────────────────────────────────────────────"];
      logs.forEach(function(log) {
        var ok = log.status >= 200 && log.status < 300;
        lines.push((ok ? "✓" : "✗") + " [" + formatTimestamp(log.timestamp) + "] "
                   + log.name + " · HTTP " + log.status
                   + (log.manual ? " （手动）" : ""));
        lines.push("  " + log.method + " " + log.url);
        if (log.response) {
          var r = log.response;
          if (r.length > 200) r = r.substring(0, 200) + "...";
          lines.push("  " + r);
        }
        lines.push("");
      });
      showResult(lines.join("\n"), false);
      document.getElementById("result").scrollIntoView({behavior: "smooth", block: "center"});
    });
}

// ---------- 启动 ----------
function onOperationChange() {
  captureSharedParams();   // 记下当前表单里的值
  renderParams();          // 渲染新操作的表单（默认值）
  restoreSharedParams();   // 把之前的值恢复回去
}

function init() {
  document.getElementById("operation").addEventListener("change", onOperationChange);
  renderParams();
  // 初始渲染后，把当前 datasheetId 记下来作为基准
  var dstEl = document.getElementById("param_datasheetId");
  if (dstEl) lastSharedDatasheetId = dstEl.value || "";
  refreshTemplateSelect();
  refreshSchedules();
  var raw = localStorage.getItem("aitable_baseconfig");
  if (raw) {
    try {
      var cfg = JSON.parse(raw);
      if (cfg.baseUrl) document.getElementById("baseUrl").value = cfg.baseUrl;
      if (cfg.token) document.getElementById("token").value = cfg.token;
    } catch (e) {}
  }
}

if (document.readyState === "loading") {
  document.addEventListener("DOMContentLoaded", init);
} else { init(); }
</script>
</body>
</html>"""


def build_request(op, base_url, datasheet_id, params):
    import urllib.parse as up

    if op == "get_spaces":
        return "GET", f"{base_url}/spaces", None

    if not datasheet_id:
        raise ValueError("datasheetId 不能为空")

    if op == "get_views":
        return "GET", f"{base_url}/datasheets/{datasheet_id}/views", None

    if op == "get_fields":
        url = f"{base_url}/datasheets/{datasheet_id}/fields"
        if params.get("viewId"):
            url += "?" + up.urlencode({"viewId": params["viewId"]})
        return "GET", url, None

    if op == "get_records":
        qp = {}
        for k in ("viewId", "pageSize", "pageNum", "maxRecords", "filterByFormula", "fieldKey"):
            if params.get(k):
                qp[k] = params[k]
        if params.get("recordIds"):
            qp["recordIds"] = [x.strip() for x in params["recordIds"].split(",") if x.strip()]
        if params.get("fields"):
            qp["fields"] = [x.strip() for x in params["fields"].split(",") if x.strip()]
        url = f"{base_url}/datasheets/{datasheet_id}/records"
        if qp:
            url += "?" + up.urlencode(qp, doseq=True)
        return "GET", url, None

    if op == "create":
        records = json.loads(params["records"])
        body = {"records": records, "fieldKey": params.get("fieldKey") or "name"}
        url = f"{base_url}/datasheets/{datasheet_id}/records"
        if params.get("viewId"):
            url += "?" + up.urlencode({"viewId": params["viewId"]})
        return "POST", url, body

    if op == "update":
        records = json.loads(params["records"])
        body = {"records": records, "fieldKey": params.get("fieldKey") or "name"}
        url = f"{base_url}/datasheets/{datasheet_id}/records"
        if params.get("viewId"):
            url += "?" + up.urlencode({"viewId": params["viewId"]})
        return "PATCH", url, body

    if op == "delete":
        rids = params.get("recordIds", "")
        if not rids:
            raise ValueError("recordIds 不能为空")
        url = f"{base_url}/datasheets/{datasheet_id}/records?" + up.urlencode({"recordIds": rids})
        return "DELETE", url, None

    raise ValueError(f"未知操作: {op}")


# ---------- 定时任务调度器 ----------
_sched_lock = threading.Lock()


def load_schedules():
    if not os.path.exists(SCHEDULES_FILE):
        return []
    try:
        with open(SCHEDULES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def save_schedules(schedules):
    with open(SCHEDULES_FILE, "w", encoding="utf-8") as f:
        json.dump(schedules, f, ensure_ascii=False, indent=2)


def load_logs():
    if not os.path.exists(LOGS_FILE):
        return []
    try:
        with open(LOGS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def save_logs(logs):
    with open(LOGS_FILE, "w", encoding="utf-8") as f:
        json.dump(logs[-LOG_LIMIT:], f, ensure_ascii=False, indent=2)


def append_log(entry):
    with _sched_lock:
        logs = load_logs()
        logs.append(entry)
        save_logs(logs)


# ---------- 文件读取（支持 CSV / XLSX） ----------
def read_table_file(path):
    """返回 (headers: list[str], rows: list[list[str]])。出错抛异常。"""
    path = os.path.expanduser(path)
    if not os.path.exists(path):
        raise FileNotFoundError(f"文件不存在: {path}")
    if not os.path.isfile(path):
        raise ValueError(f"路径不是文件: {path}")

    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            all_rows = list(reader)
        if not all_rows:
            return [], []
        headers = [str(c).strip() for c in all_rows[0]]
        data = [list(r) for r in all_rows[1:]]
        # 去掉末尾全空行
        while data and all((c == "" or c is None) for c in data[-1]):
            data.pop()
        return headers, data

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl  # type: ignore
        except ImportError:
            raise RuntimeError("读取 .xlsx 需要 openpyxl 库，请运行：pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(["" if v is None else str(v) for v in row])
        if not all_rows:
            return [], []
        headers = [str(c).strip() for c in all_rows[0]]
        data = all_rows[1:]
        while data and all((c == "" or c is None) for c in data[-1]):
            data.pop()
        return headers, data

    raise ValueError(f"不支持的文件类型：{ext}（仅支持 .csv / .xlsx / .xls）")


# ---------- 字段值类型转换（与前端一致） ----------
def coerce_field_value_for_import(value, field_type):
    """根据 AITable 字段 type，把字符串转成合适的 Python 类型。"""
    if value is None or value == "":
        return None
    s = str(value).strip()
    if field_type in ("Number", "Currency", "Percent", "Rating", "AutoNumber"):
        try:
            return float(s)
        except ValueError:
            return value
    if field_type == "Checkbox":
        lower = s.lower()
        if lower in ("true", "1", "是", "yes", "✓"):
            return True
        if lower in ("false", "0", "否", "no", ""):
            return False
        return bool(s)
    if field_type == "DateTime":
        # 尝试常见格式
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y/%m/%d %H:%M:%S"):
            try:
                dt = datetime.strptime(s, fmt)
                return int(dt.timestamp() * 1000)
            except ValueError:
                continue
        return value
    if field_type == "MultiSelect":
        import re
        return [x.strip() for x in re.split(r"[,，;；]", s) if x.strip()]
    return value


# ---------- 拉 schema（后端，给定时任务用）----------
def fetch_field_schema(token, base_url, datasheet_id):
    method, url, _ = build_request("get_fields", base_url, datasheet_id, {})
    status, resp_text = call_aitable(method, url, token, None)
    if status >= 400:
        return None
    try:
        data = json.loads(resp_text)
        return data.get("data", {}).get("fields", [])
    except Exception:
        return None


# ---------- 拉去重所需的现有值 ----------
def fetch_existing_dedup_values(token, base_url, datasheet_id, dedup_key):
    """翻页拉所有记录，提取 dedup_key 的非空值集合。"""
    existing = set()
    total_records = 0
    page = 1
    while True:
        params = {
            "fieldKey": "name",
            "fields": dedup_key,
            "pageSize": "1000",
            "pageNum": str(page),
        }
        method, url, _ = build_request("get_records", base_url, datasheet_id, params)
        status, resp_text = call_aitable(method, url, token, None)
        if status >= 400:
            raise RuntimeError(f"查现有记录失败 HTTP {status}")
        try:
            data = json.loads(resp_text)
        except Exception:
            raise RuntimeError("响应不是 JSON")
        records = (data.get("data", {}) or {}).get("records", []) or []
        total_records += len(records)
        for r in records:
            v = (r.get("fields") or {}).get(dedup_key)
            if v is not None and v != "":
                existing.add(str(v))
        if len(records) < 1000:
            break
        page += 1
    return existing, total_records


def execute_schedule(sched):
    """执行单个定时任务，返回 (status, method, url, response_text)。
    如果有 importConfig，走批量导入流程；否则走单次 API 调用。
    """
    try:
        if sched.get("importConfig"):
            return execute_import_schedule(sched)
        method, url, body = build_request(
            sched["op"], sched["baseUrl"],
            sched.get("datasheetId", ""), sched.get("params", {})
        )
        status, resp_text = call_aitable(method, url, sched["token"], body)
        return status, method, url, resp_text
    except Exception as e:
        return 500, "?", "?", f"内部错误: {e}"


def execute_import_schedule(sched):
    """从文件读 → 转换 → 去重 → 分批 POST。"""
    cfg = sched["importConfig"]
    file_path = cfg.get("filePath", "")
    if not file_path:
        return 500, "?", "?", "importConfig 缺少 filePath"

    token = sched["token"]
    base_url = sched["baseUrl"]
    datasheet_id = sched.get("datasheetId", "")
    if not datasheet_id:
        return 500, "?", "?", "缺少 datasheetId"

    # 1. 读文件
    try:
        headers, file_rows = read_table_file(file_path)
    except Exception as e:
        return 500, "?", "?", f"读文件失败: {e}"

    if not headers:
        return 500, "?", "?", "文件没有表头"

    # 2. 应用起始行偏移
    start_row = int(cfg.get("startRow", 1)) or 1
    file_rows = file_rows[start_row - 1:]

    # 3. 拉 schema 用于类型转换
    schema = fetch_field_schema(token, base_url, datasheet_id) or []
    type_by_name = {f.get("name"): f.get("type") for f in schema}

    # 4. 应用列映射 → 构造 records
    mapping = cfg.get("mapping", {})  # { "0": "字段A", "1": "字段B" }
    records = []
    for row in file_rows:
        fields = {}
        for col_idx_str, aitable_field in mapping.items():
            if not aitable_field:
                continue
            try:
                col_idx = int(col_idx_str)
            except (ValueError, TypeError):
                continue
            if col_idx >= len(row):
                continue
            raw = row[col_idx]
            if raw is None or raw == "":
                continue
            v = coerce_field_value_for_import(raw, type_by_name.get(aitable_field))
            if v is None:
                continue
            fields[aitable_field] = v
        if fields:
            records.append({"fields": fields})

    if not records:
        return 200, "POST", f"(file:{os.path.basename(file_path)})", "处理后没有有效行（可能起始行后无数据，或所有列都被忽略）"

    skipped = 0
    # 5. 去重（如果配置了）
    if cfg.get("dedup") and cfg.get("dedupKey"):
        try:
            existing_set, total_records = fetch_existing_dedup_values(
                token, base_url, datasheet_id, cfg["dedupKey"]
            )
        except Exception as e:
            return 500, "?", "?", f"去重时失败: {e}"
        before = len(records)
        records = [r for r in records if str(r["fields"].get(cfg["dedupKey"], "")) not in existing_set]
        skipped = before - len(records)
        if not records:
            return 200, "POST", f"(file:{os.path.basename(file_path)})", \
                   f"全部 {before} 条已存在（API 表共 {total_records} 条记录，{len(existing_set)} 个唯一值），无需新增"

    # 6. 分批 POST
    batch_size = 10
    success = 0
    failures = []
    last_url = ""
    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        method, url, body = build_request(
            "create", base_url, datasheet_id,
            {"records": json.dumps(batch, ensure_ascii=False), "fieldKey": "name"}
        )
        last_url = url
        status, resp_text = call_aitable(method, url, token, body)
        if 200 <= status < 300:
            success += len(batch)
        else:
            failures.append(f"批 {i // batch_size + 1}: HTTP {status} {resp_text[:120]}")

    summary = (
        f"文件 {os.path.basename(file_path)}：成功 {success}/{len(records)} 条"
        + (f"（去重跳过 {skipped} 条）" if skipped else "")
        + (("\n失败：" + "; ".join(failures)) if failures else "")
    )
    final_status = 200 if not failures else 207
    return final_status, "POST", last_url, summary


def should_fire_now(sched, now_dt, now_ms):
    """判断一个任务现在该不该跑。
    支持两种 scheduleType:
      interval - 每隔 N 分钟
      weekly   - 每周某几天的某个时点
    """
    if not sched.get("enabled"):
        return False

    sched_type = sched.get("scheduleType", "interval")

    if sched_type == "interval":
        interval_ms = int(sched.get("intervalMinutes", 60)) * 60 * 1000
        last_run = sched.get("lastRun", 0)
        return (now_ms - last_run) >= interval_ms

    if sched_type == "weekly":
        days = sched.get("daysOfWeek", [])  # [1..7] (1=周一)
        time_str = sched.get("time", "09:00")
        try:
            target_h, target_m = [int(x) for x in time_str.split(":")]
        except ValueError:
            return False
        weekday = now_dt.isoweekday()  # 1=Mon ... 7=Sun
        if weekday not in days:
            return False
        # 必须已过设定时点
        if (now_dt.hour, now_dt.minute) < (target_h, target_m):
            return False
        # 同一天只触发一次
        last_run = sched.get("lastRun", 0)
        if last_run:
            last_run_date = datetime.fromtimestamp(last_run / 1000).date()
            if last_run_date == now_dt.date():
                return False
        return True

    return False


def scheduler_loop():
    """后台线程：每 30 秒检查一次哪些任务该跑了。"""
    while True:
        try:
            time.sleep(30)
            now_dt = datetime.now()
            now_ms = int(time.time() * 1000)
            with _sched_lock:
                schedules = load_schedules()
            changed = False
            for sched in schedules:
                if not should_fire_now(sched, now_dt, now_ms):
                    continue
                status, method, url, resp_text = execute_schedule(sched)
                sched["lastRun"] = now_ms
                sched["lastStatus"] = status
                resp_short = (resp_text or "")[:600]
                append_log({
                    "scheduleId": sched["id"],
                    "name": sched.get("name", ""),
                    "timestamp": now_ms,
                    "status": status,
                    "method": method,
                    "url": url,
                    "response": resp_short,
                    "manual": False,
                })
                changed = True
            if changed:
                with _sched_lock:
                    save_schedules(schedules)
        except Exception as e:
            print("scheduler error:", e)


def start_scheduler():
    t = threading.Thread(target=scheduler_loop, daemon=True)
    t.start()


def call_aitable(method, url, token, body):
    headers = {"Authorization": f"Bearer {token}"}
    data = None
    if body is not None:
        headers["Content-Type"] = "application/json"
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=data, method=method, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return resp.status, resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8", errors="replace")


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        pass

    # ---------- 工具方法 ----------
    def _read_json(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length).decode("utf-8")
        return json.loads(raw)

    def _respond_json(self, status, data):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    # ---------- GET ----------
    def do_GET(self):
        if self.path in ("/", "/index.html"):
            body = INDEX_HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Cache-Control", "no-store")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if self.path == "/api/schedules":
            with _sched_lock:
                self._respond_json(200, load_schedules())
            return
        if self.path == "/api/schedule-logs":
            with _sched_lock:
                logs = load_logs()
            self._respond_json(200, list(reversed(logs[-100:])))
            return
        self.send_response(404); self.end_headers()

    # ---------- POST ----------
    def do_POST(self):
        if self.path == "/api/call":
            self._handle_api_call(); return
        if self.path == "/api/schedules":
            self._handle_create_schedule(); return
        if self.path.startswith("/api/schedules/run/"):
            sid = self.path[len("/api/schedules/run/"):]
            self._handle_run_now(sid); return
        if self.path == "/api/read-file":
            self._handle_read_file(); return
        self.send_response(404); self.end_headers()

    def _handle_read_file(self):
        try:
            payload = self._read_json()
            path = payload.get("path", "").strip()
            if not path:
                self._respond_json(200, {"error": "请填写文件路径"})
                return
            headers, rows = read_table_file(path)
            # 限制传给前端的行数，避免超大
            preview_rows = rows[:5000]
            self._respond_json(200, {"headers": headers, "rows": preview_rows, "totalRows": len(rows)})
        except Exception as e:
            self._respond_json(200, {"error": str(e)})

    def _handle_api_call(self):
        try:
            payload = self._read_json()
            op = payload["op"]
            token = payload["token"]
            base_url = payload["baseUrl"]
            datasheet_id = payload.get("datasheetId", "")
            params = payload.get("params", {})

            method, url, body = build_request(op, base_url, datasheet_id, params)
            status, resp_text = call_aitable(method, url, token, body)

            try:
                resp_obj = json.loads(resp_text)
            except Exception:
                resp_obj = resp_text

            self._respond_json(200, {
                "status": status,
                "request": {"method": method, "url": url, "body": body},
                "response": resp_obj,
            })
        except Exception as e:
            self._respond_json(200, {
                "status": 500,
                "request": {"method": "?", "url": "?", "body": None},
                "response": f"内部错误: {e}",
            })

    def _handle_create_schedule(self):
        try:
            payload = self._read_json()
            sched = {
                "id": uuid.uuid4().hex[:12],
                "name": payload.get("name", "未命名"),
                "scheduleType": payload.get("scheduleType", "interval"),  # interval / weekly
                "intervalMinutes": int(payload.get("intervalMinutes", 60)),
                "daysOfWeek": payload.get("daysOfWeek", []) or [],
                "time": payload.get("time", "09:00"),
                "enabled": bool(payload.get("enabled", True)),
                "createdAt": int(time.time() * 1000),
                "lastRun": 0,
                "lastStatus": None,
                "op": payload["op"],
                "token": payload["token"],
                "baseUrl": payload["baseUrl"],
                "datasheetId": payload.get("datasheetId", ""),
                "params": payload.get("params", {}),
                "importConfig": payload.get("importConfig"),  # None 或 {filePath, mapping, startRow, dedup, dedupKey}
            }
            with _sched_lock:
                schedules = load_schedules()
                schedules.append(sched)
                save_schedules(schedules)
            self._respond_json(200, sched)
        except Exception as e:
            self._respond_json(400, {"error": str(e)})

    def _handle_run_now(self, sid):
        with _sched_lock:
            schedules = load_schedules()
        target = None
        for s in schedules:
            if s["id"] == sid:
                target = s
                break
        if not target:
            self._respond_json(404, {"error": "schedule not found"}); return
        status, method, url, resp_text = execute_schedule(target)
        now_ms = int(time.time() * 1000)
        target["lastRun"] = now_ms
        target["lastStatus"] = status
        with _sched_lock:
            save_schedules(schedules)
        append_log({
            "scheduleId": sid, "name": target.get("name", ""),
            "timestamp": now_ms, "status": status,
            "method": method, "url": url,
            "response": (resp_text or "")[:600], "manual": True,
        })
        self._respond_json(200, {"status": status, "response": resp_text})

    # ---------- DELETE / PATCH ----------
    def do_DELETE(self):
        if self.path.startswith("/api/schedules/"):
            sid = self.path[len("/api/schedules/"):]
            with _sched_lock:
                schedules = load_schedules()
                new_list = [s for s in schedules if s["id"] != sid]
                save_schedules(new_list)
            self._respond_json(200, {"ok": True}); return
        self.send_response(404); self.end_headers()

    def do_PATCH(self):
        if self.path.startswith("/api/schedules/"):
            sid = self.path[len("/api/schedules/"):]
            try:
                payload = self._read_json()
            except Exception as e:
                self._respond_json(400, {"error": str(e)}); return
            with _sched_lock:
                schedules = load_schedules()
                for s in schedules:
                    if s["id"] == sid:
                        if "enabled" in payload: s["enabled"] = bool(payload["enabled"])
                        if "intervalMinutes" in payload: s["intervalMinutes"] = int(payload["intervalMinutes"])
                        if "name" in payload: s["name"] = payload["name"]
                        if "scheduleType" in payload: s["scheduleType"] = payload["scheduleType"]
                        if "daysOfWeek" in payload: s["daysOfWeek"] = payload["daysOfWeek"] or []
                        if "time" in payload: s["time"] = payload["time"]
                        save_schedules(schedules)
                        self._respond_json(200, s); return
            self._respond_json(404, {"error": "not found"}); return
        self.send_response(404); self.end_headers()


def main():
    start_scheduler()
    server = HTTPServer(("127.0.0.1", PORT), Handler)
    url = f"http://127.0.0.1:{PORT}/"
    print(f"\n  AITable 调试工具已启动")
    print(f"  访问: {url}")
    print(f"  定时任务调度器已运行（仅在工具运行期间生效）")
    print(f"  按 Ctrl+C 退出\n")
    try: webbrowser.open(url)
    except Exception: pass
    try: server.serve_forever()
    except KeyboardInterrupt: print("\n已退出")


if __name__ == "__main__":
    main()
